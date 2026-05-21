"""Command-line interface for the leakage clusterer.

``python -m proteosphere overlap-cluster --input accessions.csv --out leakage.json``

Reads a flat list of UniProt accessions, computes leakage clusters under
the chosen severity tiers, and writes a JSON manifest the splitter can
consume.

Two usage modes:

1. **Non-interactive** (good for scripts and CI):
   pass ``--tiers`` to choose constraints up front and ``--out`` to dump
   the manifest in one shot.

2. **Interactive** (good for first-time exploration):
   pass ``--interactive``. The CLI runs once with the default tier set,
   shows what fires, and lets you add/remove tiers and re-run until
   you're happy. Press ``w`` to write the manifest.
"""
from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path

from proteosphere.config import Config

from .clusters import (
    DEFAULT_CLUSTER_TIERS,
    DEFAULT_MIN_SPECIFICITY,
    LeakageManifest,
    TIER_SOURCES,
    compute_leakage_clusters,
)
from .column_aliases import (
    ACCESSION_COLUMNS,
    COMPARISON_COLUMNS,
    PAIR_A_COLUMNS,
    PAIR_B_COLUMNS,
    PDB_COLUMNS,
    PROTEIN_COLUMNS,
    ROW_ID_COLUMNS,
    TEST_COLUMNS,
    find_all_column_indices,
    looks_like_header_row,
    normalize as _normalize_col,
)
# Reuse painter / terminal helpers from the scan CLI so the two commands
# look and feel identical.
from .cli import (
    MIN_TERM_WIDTH,
    _Ansi,
    _Painter,
    _enable_ansi_on_windows,
    _fmt_duration,
    _supports_color,
    _terminal_width,
    _truncate,
    _resolve_pdb_set,
    _TIER_STYLE,
    resolve_config_or_friendly_exit,
    resolve_pdb_details,
)
from .tiers import SEVERITY_TIERS, TIER_DESCRIPTIONS


# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------

# Tiers that aren't useful for clustering — show them in --list-tiers but
# don't include in default and warn if user opts in.
_NON_CLUSTERABLE_TIERS = frozenset({"shared_partial_architecture"})

# Display column widths.
_LABEL_COL_WIDTH = 28

# Heartbeat tick rate for the live progress display (seconds). Fast
# enough that the spinner stays lively, slow enough to keep cost trivial.
_HEARTBEAT_INTERVAL = 0.25
_SPINNER_FRAMES = "|/-\\"
# Total progress-bar width (characters of fill area, excluding the
# brackets).
_PROGRESS_BAR_WIDTH = 20


@dataclass
class _ClusterProgress:
    """Live single-line progress display for the clusterer.

    Renders one line with an overall progress bar, the current source
    being queried, the spinner, and elapsed time. On a TTY the line
    updates in place via ``\\r``; on a non-TTY (redirected to a file or
    pipe) we emit one completion line per source so logs stay clean.
    """
    painter: _Painter
    tty: bool
    total_sources: int = 0

    _lock: threading.Lock = field(default_factory=threading.Lock)
    _stop: threading.Event = field(default_factory=threading.Event)
    _thread: threading.Thread | None = None
    _tick: int = 0

    # Per-source state
    _current_tier: str = ""
    _current_table: str = ""
    _current_namespace: str = ""
    _current_index: int = 0
    _source_start: float = 0.0

    # Cumulative
    _start_time: float = 0.0
    _completed_sources: int = 0
    _total_union_events: int = 0

    # ----- thread lifecycle ------------------------------------------

    def start(self, total_sources: int) -> None:
        self.total_sources = total_sources
        self._start_time = time.monotonic()
        if not self.tty:
            return
        self._thread = threading.Thread(
            target=self._heartbeat_loop, daemon=True,
        )
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=1.0)
            self._thread = None

    def _heartbeat_loop(self) -> None:
        while not self._stop.wait(_HEARTBEAT_INTERVAL):
            with self._lock:
                self._tick += 1
                if self._current_tier and self.tty:
                    self._paint_inplace()

    # ----- engine callbacks -----------------------------------------

    def on_source_start(
        self, tier: str, table: str, namespace: str,
        i: int, total: int,
    ) -> None:
        with self._lock:
            self._current_tier = tier
            self._current_table = table
            self._current_namespace = namespace
            self._current_index = i
            self._source_start = time.monotonic()
            if self.tty:
                self._paint_inplace()

    def on_source_done(
        self, tier: str, table: str, namespace: str,
        i: int, total: int, n_union_events: int,
    ) -> None:
        with self._lock:
            self._completed_sources = i
            self._total_union_events += n_union_events
            source_elapsed = time.monotonic() - self._source_start
            if self.tty:
                # Clear the in-place line; we'll repaint with the new
                # source on the next on_source_start.
                self._paint_inplace()
            else:
                # Non-TTY: emit one final line per source.
                line = (
                    f"  [{i:>3}/{total:>3}]  "
                    f"{tier}/{namespace:<22}  "
                    f"{_fmt_duration(source_elapsed):>6}  "
                    f"{n_union_events:>3} union event(s)"
                )
                sys.stdout.write(line + "\n")
                sys.stdout.flush()

    def finish(self) -> None:
        """Render a final summary line after everything's done."""
        elapsed = time.monotonic() - self._start_time
        with self._lock:
            if self.tty:
                # Clear the in-place line.
                cols = _terminal_width()
                sys.stdout.write("\r" + " " * (cols - 1) + "\r")
            sys.stdout.write(self.painter.dim(
                f"  Done. {self._completed_sources}/{self.total_sources} sources, "
                f"{self._total_union_events} union events, "
                f"in {_fmt_duration(elapsed)}.\n"
            ))
            sys.stdout.flush()

    # ----- rendering -------------------------------------------------

    def _paint_inplace(self) -> None:
        cols = _terminal_width()
        # Progress bar
        if self.total_sources > 0:
            done = self._current_index - 1  # current is in-progress
            done = max(0, done)
            pct = done / self.total_sources
        else:
            pct = 0.0
        filled = int(round(pct * _PROGRESS_BAR_WIDTH))
        bar = "[" + "#" * filled + "-" * (_PROGRESS_BAR_WIDTH - filled) + "]"
        pct_text = f"{pct * 100:>3.0f}%"

        spinner = _SPINNER_FRAMES[self._tick % len(_SPINNER_FRAMES)]
        source_elapsed = time.monotonic() - self._source_start
        source_str = (
            f"{self._current_tier}/{self._current_namespace}"
            if self._current_tier else ""
        )
        total_elapsed = time.monotonic() - self._start_time

        # Build a single line, truncated to terminal width.
        line = (
            f"  {bar} {pct_text}  "
            f"({self._current_index}/{self.total_sources})  "
            f"{source_str} {spinner}  "
            f"T+{_fmt_duration(total_elapsed)}"
        )
        # Pad / trim to terminal width to overwrite previous render.
        if len(line) >= cols:
            line = line[: cols - 1]
        else:
            line = line + " " * (cols - len(line) - 1)
        sys.stdout.write("\r" + line)
        sys.stdout.flush()


# ---------------------------------------------------------------------------
# Input loading: accept either a flat accession list or a pairs file
# ---------------------------------------------------------------------------


def _split_id_cell(cell: str) -> list[str]:
    """Split a CSV cell that may contain multiple IDs joined by ; | ,"""
    if not cell:
        return []
    text = cell.strip()
    if not text:
        return []
    for sep in (";", "|"):
        if sep in text:
            return [t.strip() for t in text.split(sep) if t.strip()]
    if "," in text:
        return [t.strip() for t in text.split(",") if t.strip()]
    return [text]


# The clusterer accepts cells from any column that looks like a protein
# / structure identifier. We sweep accession + PDB + protein + test +
# comparison + pair_a + pair_b + row_id columns, which between them
# cover the realistic "tell me which column has the IDs" cases.
_ID_BEARING_COLUMNS = (
    ACCESSION_COLUMNS | PDB_COLUMNS | PROTEIN_COLUMNS
    | TEST_COLUMNS | COMPARISON_COLUMNS
    | PAIR_A_COLUMNS | PAIR_B_COLUMNS
    | ROW_ID_COLUMNS
)


def _looks_like_header(row: list[str]) -> bool:
    """Forward to the centralized detector."""
    return looks_like_header_row(row)


def _looks_like_id(token: str) -> bool:
    """Return True if ``token`` looks like a UniProt accession or PDB ID.

    Used by the column-sweep loader to ignore non-ID columns (scores,
    features, descriptions, ...). UniProt accessions are 6-10 chars,
    start with a letter, all alphanumeric. PDB IDs are exactly 4 chars,
    start with a digit, all alphanumeric.
    """
    t = (token or "").strip()
    if not t:
        return False
    # PDB: 4 chars, digit-first, alnum.
    if len(t) == 4 and t[0].isdigit() and t.isalnum():
        return True
    # UniProt accession: 6-10 chars, letter-first, alnum, contains at
    # least one digit (excludes things like "ALPHA", "BETA", "TRAIN").
    if 6 <= len(t) <= 10 and t[0].isalpha() and t.isalnum() and any(c.isdigit() for c in t):
        return True
    return False


def _load_csv_accessions(path: Path, column_hint: str | None = None) -> list[str]:
    """Read a CSV/TSV and return every unique ID it contains.

    The loader is permissive:
    - If a column matches the hint (or a built-in alias for the
      accession / PDB column), it reads only that column from data rows.
    - Otherwise, it sweeps every data row's every cell — skipping the
      first row when it looks like a header row.

    Header detection avoids the common bug of accidentally treating
    'pdb_id' (the column name) as an ID.
    """
    delim = "\t" if path.suffix.lower() in (".tsv", ".txt") else ","
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=delim)
        rows = list(reader)
    if not rows:
        return []
    header_row = [h if h else "" for h in rows[0]]
    has_header = _looks_like_header(header_row)

    # We accept multiple ID-bearing columns. If the user passed
    # ``--column foo,bar``, those are tried first (exact normalized
    # match against the header). Otherwise we sweep every column whose
    # header name matches any of the known ID-bearing roles.
    target_cols: list[int] = []
    if column_hint:
        wanted = {_normalize_col(w) for w in column_hint.split(",") if w.strip()}
        for i, name in enumerate(header_row):
            if _normalize_col(name) in wanted:
                target_cols.append(i)
    if not target_cols and has_header:
        target_cols = find_all_column_indices(header_row, _ID_BEARING_COLUMNS)

    data_rows = rows[1:] if has_header else rows

    seen: dict[str, None] = {}
    if target_cols:
        # User pointed us at a column (or we matched on a header alias).
        # Trust every non-empty cell in that column.
        for row in data_rows:
            for col in target_cols:
                if col >= len(row):
                    continue
                for tok in _split_id_cell(row[col]):
                    seen[tok] = None
    else:
        # No known column name — sweep every cell, but FILTER to tokens
        # that actually look like IDs. Without this we'd pick up scores,
        # feature columns, organism names, etc.
        for row in data_rows:
            for cell in row:
                for tok in _split_id_cell(cell or ""):
                    if _looks_like_id(tok):
                        seen[tok] = None
    return list(seen)


def _load_json_accessions(path: Path) -> list[str]:
    """Read JSON of various shapes and extract a flat accession list.

    Accepted shapes:
        ["P00001", "P00002"]
        {"accessions": ["P00001", ...]}
        {"pairs": [{"test": [...], "comparison": [...]}, ...]}
    """
    payload = json.loads(path.read_text(encoding="utf-8"))
    seen: dict[str, None] = {}
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, str):
                for tok in _split_id_cell(item):
                    seen[tok] = None
            elif isinstance(item, dict):
                for key in ("accession", "test", "comparison", "a", "b"):
                    val = item.get(key)
                    if isinstance(val, list):
                        for v in val:
                            if v is not None:
                                for tok in _split_id_cell(str(v)):
                                    seen[tok] = None
                    elif isinstance(val, str):
                        for tok in _split_id_cell(val):
                            seen[tok] = None
    elif isinstance(payload, dict):
        if "accessions" in payload and isinstance(payload["accessions"], list):
            for v in payload["accessions"]:
                if v is not None:
                    for tok in _split_id_cell(str(v)):
                        seen[tok] = None
        if "pairs" in payload and isinstance(payload["pairs"], list):
            for item in payload["pairs"]:
                if not isinstance(item, dict):
                    continue
                for key in ("test", "comparison", "a", "b"):
                    val = item.get(key)
                    if isinstance(val, list):
                        for v in val:
                            if v is not None:
                                for tok in _split_id_cell(str(v)):
                                    seen[tok] = None
                    elif isinstance(val, str):
                        for tok in _split_id_cell(val):
                            seen[tok] = None
    return list(seen)


def _load_xlsx_accessions(path: Path, column_hint: str | None = None) -> list[str]:
    """Read an .xlsx file's first sheet and return every ID it contains.

    Mirrors the CSV loader's behavior:
    - Detects a header row via :func:`_looks_like_header` so column names
      (``train_ids``, ``test_ids``, etc.) aren't mistaken for IDs.
    - Picks a column by hint or by matching :data:`_COLUMN_ALIASES`.
    - Falls back to sweeping every cell, filtered to tokens that actually
      look like UniProt accessions or PDB IDs.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install with:\n"
            "  pip install openpyxl\n"
            "or:  pip install proteosphere[xlsx]"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_raw:
        return []

    # Normalise to strings (XLSX cells can be int/float/None).
    rows: list[list[str]] = [
        ["" if c is None else str(c) for c in row] for row in rows_raw
    ]
    header_row = [h if h else "" for h in rows[0]]
    has_header = _looks_like_header(header_row)

    target_cols: list[int] = []
    if column_hint:
        wanted = {_normalize_col(w) for w in column_hint.split(",") if w.strip()}
        for i, name in enumerate(header_row):
            if _normalize_col(name) in wanted:
                target_cols.append(i)
    if not target_cols and has_header:
        target_cols = find_all_column_indices(header_row, _ID_BEARING_COLUMNS)

    data_rows = rows[1:] if has_header else rows
    seen: dict[str, None] = {}

    if target_cols:
        for row in data_rows:
            for col in target_cols:
                if col >= len(row):
                    continue
                for tok in _split_id_cell(row[col]):
                    seen[tok] = None
    else:
        # No known column name — sweep every cell, filtered to tokens
        # that actually look like IDs.
        for row in data_rows:
            for cell in row:
                for tok in _split_id_cell(cell or ""):
                    if _looks_like_id(tok):
                        seen[tok] = None
    return list(seen)


def load_accessions(path: Path, *, column_hint: str | None = None) -> list[str]:
    """Dispatch on file extension."""
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        return _load_csv_accessions(path, column_hint=column_hint)
    if suffix == ".json":
        return _load_json_accessions(path)
    if suffix == ".xlsx":
        return _load_xlsx_accessions(path, column_hint=column_hint)
    raise SystemExit(
        f"Unsupported input format: {suffix!r}. Supported: .csv .tsv .json .xlsx"
    )


def _looks_like_pdb(token: str) -> bool:
    """Return True if ``token`` looks like a 4-character PDB ID."""
    return len(token) == 4 and token[0].isdigit() and token.isalnum()


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------


@dataclass
class _ClusterSummary:
    """Compact view of a manifest used for tabular display."""
    cluster_count: int
    singleton_count: int
    clustered_accession_count: int
    tier_counts: dict[str, int]  # tier -> count of clusters with that tier
    largest_cluster_size: int


def _summarise(manifest: LeakageManifest) -> _ClusterSummary:
    if manifest.clusters:
        largest = max(len(c.members) for c in manifest.clusters)
    else:
        largest = 0
    return _ClusterSummary(
        cluster_count=len(manifest.clusters),
        singleton_count=len(manifest.singletons),
        clustered_accession_count=sum(len(c.members) for c in manifest.clusters),
        tier_counts=manifest.tier_counts(),
        largest_cluster_size=largest,
    )


def _render_header(
    *, painter: _Painter, input_path: Path, n_accs: int,
    tiers: list[str], min_specificity: int | None,
    warehouse_root: Path, width: int,
) -> None:
    title = "ProteoSphere Leakage Clusterer"
    print(painter.bold(title))
    print("=" * len(title))
    print()
    fields = [
        ("Input          ", f"{input_path}  ({n_accs} unique accession{'s' if n_accs != 1 else ''})"),
        ("Tiers          ", ", ".join(tiers)),
        ("Min specificity",
         f"{min_specificity} worldwide members" if min_specificity is not None else "no cap"),
        ("Warehouse      ", str(warehouse_root)),
    ]
    for name, value in fields:
        line = f"  {name}  {value}"
        if len(line) > width:
            line = line[: width - 3] + "..."
        print(line)
    print()


def _render_summary(
    *, painter: _Painter, summary: _ClusterSummary, manifest: LeakageManifest, width: int,
) -> None:
    n_input = len(manifest.input_accessions)
    print(painter.bold("Summary"))
    print("-------")
    print(f"  Input accessions       : {n_input}")
    print(f"  Clusters (>=2 members) : {summary.cluster_count}")
    print(f"  Clustered accessions   : {summary.clustered_accession_count}")
    print(f"  Singletons             : {summary.singleton_count}")
    if summary.largest_cluster_size:
        print(f"  Largest cluster        : {summary.largest_cluster_size} accessions")
    if summary.tier_counts:
        print()
        print(painter.bold("  Clusters by contributing tier"))
        for tier in SEVERITY_TIERS:
            n = summary.tier_counts.get(tier, 0)
            if n == 0:
                continue
            color, marker = _TIER_STYLE.get(tier, (_Ansi.DIM, "   "))
            line = f"    {tier:<28} {n:>4}   {marker}"
            print(painter.tier(tier, line))


def _render_clusters(
    *, painter: _Painter, manifest: LeakageManifest, max_show: int, width: int,
) -> None:
    if not manifest.clusters:
        print(painter.dim("  (no clusters found at the chosen tiers / specificity)"))
        return
    print()
    print(painter.bold(f"Clusters ({len(manifest.clusters)} total, showing up to {max_show})"))
    print("-" * 78)
    member_col = max(20, width - 52)
    for cluster in manifest.clusters[:max_show]:
        members_str = _truncate(", ".join(cluster.members), member_col)
        # Tier of the strongest source (first one — already sorted by n_bridged desc).
        top_source = cluster.sources[0] if cluster.sources else None
        top_tier = top_source.tier if top_source else "?"
        tier_text = painter.tier(top_tier, f"{top_tier:<22}")
        head = f"  {cluster.cluster_id}  size={len(cluster.members):>3}  {tier_text}"
        print(f"{head}  members: {members_str}")
        if top_source:
            extra = ""
            if len(cluster.sources) > 1:
                extra = f"  (+{len(cluster.sources)-1} more source{'s' if len(cluster.sources)>2 else ''})"
            via = (
                f"{top_source.tier}/{top_source.namespace}:{top_source.identifier}"
                + (f" [worldwide={top_source.prevalence}]" if top_source.prevalence else "")
            )
            print(painter.dim(f"      via: {via}{extra}"))
    if len(manifest.clusters) > max_show:
        print(painter.dim(
            f"  ... {len(manifest.clusters) - max_show} more cluster(s); see --out JSON for the full list."
        ))


# ---------------------------------------------------------------------------
# Interactive tier-toggling loop
# ---------------------------------------------------------------------------


def _run_with_progress(
    *, painter: _Painter, config: Config, accessions: list[str],
    tiers: list[str], min_specificity: int | None,
) -> LeakageManifest:
    """Compute a leakage manifest with live progress."""
    is_tty = sys.stdout.isatty()
    progress = _ClusterProgress(painter=painter, tty=is_tty)

    # The clusterer needs to know total_sources before the first source
    # starts so the bar can render correctly from the get-go. We
    # capture it from the first on_source_start call.
    total_holder = {"total": 0}

    def _start(tier, table, ns, i, total):
        if total_holder["total"] == 0:
            total_holder["total"] = total
            progress.start(total)
        progress.on_source_start(tier, table, ns, i, total)

    manifest = compute_leakage_clusters(
        config, accessions, tiers=tiers, min_specificity=min_specificity,
        on_source_start=_start,
        on_source_done=progress.on_source_done,
    )
    progress.stop()
    if total_holder["total"]:
        progress.finish()
    return manifest


def _interactive_loop(
    *, painter: _Painter, config: Config, accessions: list[str],
    initial_tiers: list[str], min_specificity: int | None, width: int,
) -> tuple[LeakageManifest, list[str]]:
    """Run the engine, show results, let the user toggle tiers, repeat.

    Returns the final manifest and the final tier list once the user
    types ``w`` (write).
    """
    tiers = list(initial_tiers)
    while True:
        print()
        print(painter.bold(f"Computing clusters with tiers: {tiers}"))
        manifest = _run_with_progress(
            painter=painter, config=config, accessions=accessions,
            tiers=tiers, min_specificity=min_specificity,
        )
        summary = _summarise(manifest)
        print()
        _render_summary(painter=painter, summary=summary, manifest=manifest, width=width)
        _render_clusters(painter=painter, manifest=manifest, max_show=10, width=width)
        print()
        # Interactive menu
        print(painter.bold("Tier toggles (select to include or exclude):"))
        all_tiers = list(SEVERITY_TIERS)
        tier_state = {t: (t in tiers) for t in all_tiers}
        for i, tier in enumerate(all_tiers, 1):
            mark = "[x]" if tier_state[tier] else "[ ]"
            sources = len(TIER_SOURCES.get(tier, []))
            note = ""
            if tier in _NON_CLUSTERABLE_TIERS:
                note = painter.dim("  (not cluster-friendly)")
            elif tier == "identity":
                note = painter.dim("  (always-on dedup)")
            elif sources == 0 and tier != "domain_architecture":
                note = painter.dim("  (no clustering sources)")
            line = f"  {mark}  {i:>2}. {tier:<30}{note}"
            if tier_state[tier]:
                print(painter.tier(tier, line))
            else:
                print(painter.dim(line))
        print()
        print(painter.bold("Commands:"))
        print("  <digit>   toggle tier on/off")
        print("  s <int>   set min-specificity cap (current: "
              + (f"{min_specificity}" if min_specificity else "no cap") + ")")
        print("  w         write the manifest and exit")
        print("  q         quit without writing")
        try:
            reply = input("> ").strip()
        except EOFError:
            reply = "q"
        if not reply:
            continue
        if reply.lower() == "q":
            print(painter.dim("Aborted; manifest not written."))
            sys.exit(0)
        if reply.lower() == "w":
            return manifest, tiers
        if reply.lower().startswith("s "):
            try:
                new_cap = int(reply.split(maxsplit=1)[1])
                min_specificity = new_cap if new_cap > 0 else None
                print(painter.dim(f"  min-specificity now {min_specificity}"))
            except (ValueError, IndexError):
                print(painter.dim("  (couldn't parse number; ignored)"))
            continue
        # Digit toggle
        try:
            idx = int(reply)
        except ValueError:
            print(painter.dim(f"  (didn't understand {reply!r})"))
            continue
        if not 1 <= idx <= len(all_tiers):
            print(painter.dim(f"  (digit out of range; pick 1-{len(all_tiers)})"))
            continue
        tier = all_tiers[idx - 1]
        if tier in tiers:
            tiers.remove(tier)
        else:
            tiers.append(tier)


# ---------------------------------------------------------------------------
# Argparse + main
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="proteosphere overlap-cluster",
        description=(
            "Build a leakage-cluster manifest from a flat list of UniProt "
            "accessions. The output JSON tells the splitter which accessions "
            "must stay together when partitioning train / val / test."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  proteosphere overlap-cluster --input accessions.csv --out leakage.json\n"
            "  proteosphere overlap-cluster --input pairs.csv --tiers identity,direct_ortholog,paralog_family\n"
            "  proteosphere overlap-cluster --input accessions.csv --interactive --out leakage.json\n"
            "  proteosphere overlap-cluster --list-tiers\n"
        ),
    )
    parser.add_argument(
        "--input", "-i", type=Path,
        help="CSV / TSV / JSON / XLSX file containing accessions (or a pairs file).",
    )
    parser.add_argument(
        "--column", default=None,
        help="Column name to read accessions from when the loader can't auto-detect.",
    )
    parser.add_argument(
        "--id-kind", default="accession", choices=("accession", "pdb"),
        help=(
            "How to interpret the IDs in the input file. 'accession' (default) "
            "treats them as UniProt accessions; 'pdb' resolves PDB IDs to "
            "UniProt accessions via the structure_units table before clustering."
        ),
    )
    parser.add_argument(
        "--tiers", default=",".join(DEFAULT_CLUSTER_TIERS),
        help=(
            "Comma-separated severity tiers to use as clustering constraints. "
            f"Default: {','.join(DEFAULT_CLUSTER_TIERS)}"
        ),
    )
    parser.add_argument(
        "--include-tier", action="append", default=[],
        help="Additive tier (can be repeated); merged with --tiers.",
    )
    parser.add_argument(
        "--exclude-tier", action="append", default=[],
        help="Subtractive tier (can be repeated); removed from --tiers.",
    )
    parser.add_argument(
        "--min-specificity", type=int, default=DEFAULT_MIN_SPECIFICITY,
        help=(
            "Drop clustering identifiers whose worldwide member count exceeds "
            f"this cap (default: {DEFAULT_MIN_SPECIFICITY}). Pass 0 to disable."
        ),
    )
    parser.add_argument(
        "--out", "-o", type=Path, default=None,
        help="Write the leakage manifest JSON to this path.",
    )
    parser.add_argument(
        "--interactive", action="store_true",
        help="Show clusters, toggle tiers, re-run until satisfied, then write.",
    )
    parser.add_argument(
        "--max-show", type=int, default=20,
        help="Maximum number of clusters to display in the on-screen table (default 20).",
    )
    parser.add_argument(
        "--quiet", "-q", action="store_true",
        help="Skip the cluster table; print only the summary.",
    )
    parser.add_argument(
        "--no-color", action="store_true",
        help="Disable ANSI color output.",
    )
    parser.add_argument(
        "--list-tiers", action="store_true",
        help="List the available severity tiers and exit.",
    )
    return parser


def _resolve_tiers(args: argparse.Namespace) -> list[str]:
    base = [t.strip() for t in args.tiers.split(",") if t.strip()]
    for t in args.include_tier:
        if t not in base:
            base.append(t)
    for t in args.exclude_tier:
        while t in base:
            base.remove(t)
    valid = set(TIER_SOURCES)
    unknown = [t for t in base if t not in valid]
    if unknown:
        raise SystemExit(
            f"Unknown tier(s): {unknown}\nValid tiers: {sorted(valid)}"
        )
    return base


def main(argv: list[str] | None = None, *, config: Config | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    use_color = not args.no_color and _supports_color()
    if use_color:
        _enable_ansi_on_windows()
    painter = _Painter(enabled=use_color)

    if args.list_tiers:
        print(painter.bold("Available severity tiers"))
        print("-" * 24)
        for tier in SEVERITY_TIERS:
            sources = TIER_SOURCES.get(tier, [])
            n_src = len(sources)
            note = ""
            if tier in _NON_CLUSTERABLE_TIERS:
                note = "  (not cluster-friendly)"
            elif tier == "identity":
                note = "  (always-on accession dedup)"
            elif tier == "domain_architecture":
                note = "  (exact Pfam set match)"
            elif n_src == 0:
                note = "  (no sources configured)"
            else:
                note = f"  ({n_src} source{'s' if n_src != 1 else ''})"
            print(f"  {tier:<30}{note}")
            desc = TIER_DESCRIPTIONS.get(tier, "")
            if desc:
                first_line = desc.split("\n")[0]
                print(painter.dim("    " + first_line))
        return 0

    if args.input is None:
        parser.error("--input is required (or pass --list-tiers).")

    input_path: Path = args.input.expanduser().resolve()
    if not input_path.is_file():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        return 2

    try:
        accessions = load_accessions(input_path, column_hint=args.column)
    except (ValueError, json.JSONDecodeError) as e:
        print(f"ERROR: failed to read {input_path}:\n  {e}", file=sys.stderr)
        return 2
    if not accessions:
        print(f"WARNING: no accessions found in {input_path}.", file=sys.stderr)
        return 0

    # Heuristic: auto-promote to --id-kind pdb when the input is clearly
    # PDB-dominant. We accept some junk (other column values) by using
    # majority rather than all-or-nothing. UniProt accessions are 6-10
    # chars and PDB IDs are 4 chars + digit-start, so there's no
    # ambiguity for the cells we care about.
    sample = accessions[: min(50, len(accessions))]
    n_pdb = sum(1 for a in sample if _looks_like_pdb(a))
    pdb_fraction = n_pdb / len(sample) if sample else 0.0
    if pdb_fraction >= 0.5 and args.id_kind == "accession":
        print(painter.bold(
            f"Input looks PDB-dominant ({n_pdb}/{len(sample)} of the first "
            f"{len(sample)} tokens are 4-char codes starting with a digit). "
            "Switching to --id-kind pdb."
        ))
        print(painter.dim(
            "  (pass --id-kind accession explicitly to override; pass --column "
            "to restrict the loader to a specific column.)"
        ))
        args.id_kind = "pdb"

    config = resolve_config_or_friendly_exit(config, painter=painter)

    # Resolve PDB IDs to UniProt accessions if requested.
    if args.id_kind == "pdb":
        n_input = len(accessions)
        print(painter.dim(
            f"Resolving {n_input} PDB IDs to UniProt accessions..."
        ))
        details = resolve_pdb_details(config, accessions)

        resolved: list[str] = []
        redirects: list[tuple[str, str]] = []
        unresolved_details: list[tuple[str, str, str]] = []  # (pdb, category, desc)
        unresolved_unknown: list[str] = []
        supplement_used = 0

        for d in details:
            if d.redirected_from:
                redirects.append((d.redirected_from, d.effective_pdb))
            if "supplement" in d.sources:
                supplement_used += 1
            if d.resolved:
                resolved.extend(d.uniprots)
            else:
                if d.unresolvable_reason:
                    unresolved_details.append((
                        d.requested_pdb,
                        d.unresolvable_reason.category,
                        d.unresolvable_reason.description,
                    ))
                else:
                    unresolved_unknown.append(d.requested_pdb)

        # Surface redirects transparently so the user knows we substituted.
        if redirects:
            print(painter.dim(
                f"  redirected {len(redirects)} obsolete PDB ID(s) to their "
                f"successors (per RCSB holdings-removed):"
            ))
            for old, new in redirects[:5]:
                print(painter.dim(f"    {old} -> {new}"))
            if len(redirects) > 5:
                print(painter.dim(f"    ... and {len(redirects)-5} more"))

        # Note supplement hits (informational).
        if supplement_used:
            print(painter.dim(
                f"  picked up {supplement_used} PDB(s) from the local "
                f"resolution supplement (entries missing from SIFTS)."
            ))

        # Categorised unresolved list with reasons.
        if unresolved_details or unresolved_unknown:
            n_unres = len(unresolved_details) + len(unresolved_unknown)
            print(painter.dim(
                f"  warning: {n_unres} PDB ID(s) did not resolve to any "
                f"UniProt accession; dropping them."
            ))
            if unresolved_details:
                by_cat: dict[str, list[str]] = {}
                for pdb, cat, _desc in unresolved_details:
                    by_cat.setdefault(cat, []).append(pdb)
                for cat, ids in sorted(by_cat.items()):
                    sample = ", ".join(ids[:4])
                    extra = "" if len(ids) <= 4 else f", +{len(ids)-4} more"
                    print(painter.dim(
                        f"    [{cat}] {len(ids)}: {sample}{extra}"
                    ))
            if unresolved_unknown:
                sample = ", ".join(unresolved_unknown[:4])
                extra = ("" if len(unresolved_unknown) <= 4
                         else f", +{len(unresolved_unknown)-4} more")
                print(painter.dim(
                    f"    [unknown] {len(unresolved_unknown)}: {sample}{extra}"
                ))
            print(painter.dim(
                "  Categories: 'antibody_fab' = Fab/Fv (V-region not in "
                "UniProt), 'designed' = computational design, 'chimera' = "
                "engineered fusion, 'unknown' = needs investigation."
            ))

        # Dedup but keep deterministic order.
        seen: dict[str, None] = {}
        for acc in resolved:
            seen[acc] = None
        accessions = list(seen)
        n_resolved_pdbs = sum(1 for d in details if d.resolved)
        if accessions and n_resolved_pdbs > 0:
            ratio = len(accessions) / n_resolved_pdbs
            note = (
                f"  resolved to {len(accessions)} unique UniProt accession(s) "
                f"from {n_resolved_pdbs} PDB(s)."
            )
            if ratio > 1.05:
                note += (
                    f"\n  (avg {ratio:.2f} chains/structure -- antibody Fabs "
                    f"contribute heavy + light chain; many enzymes have "
                    f"multiple subunits.)"
                )
            print(painter.dim(note))
        else:
            print(painter.dim(
                f"  resolved to {len(accessions)} unique UniProt accession(s)."
            ))
        if not accessions:
            print(painter.bold(
                "ERROR: no UniProt accessions could be resolved from the input "
                "PDB IDs. Nothing to cluster."
            ))
            return 2

    tiers = _resolve_tiers(args)
    min_specificity = None if args.min_specificity <= 0 else args.min_specificity
    width = _terminal_width()

    _render_header(
        painter=painter, input_path=input_path, n_accs=len(accessions),
        tiers=tiers, min_specificity=min_specificity,
        warehouse_root=config.warehouse_root, width=width,
    )

    if args.interactive:
        manifest, tiers = _interactive_loop(
            painter=painter, config=config, accessions=accessions,
            initial_tiers=tiers, min_specificity=min_specificity, width=width,
        )
    else:
        print(painter.dim(
            "Computing clusters... (the first source pays the cold-cache cost; "
            "subsequent sources should be much faster)"
        ))
        manifest = _run_with_progress(
            painter=painter, config=config, accessions=accessions,
            tiers=tiers, min_specificity=min_specificity,
        )
        print()
        summary = _summarise(manifest)
        _render_summary(painter=painter, summary=summary, manifest=manifest, width=width)
        if not args.quiet:
            _render_clusters(
                painter=painter, manifest=manifest, max_show=args.max_show, width=width,
            )

    if args.out is not None:
        out_path: Path = args.out.expanduser().resolve()
        manifest.write_json(out_path)
        print()
        print(f"  Wrote leakage manifest -> {out_path}")
        print(painter.dim(
            f"    Feed it to the splitter via: "
            f"proteosphere-split ... --leakage-manifest {out_path.name}"
        ))
    elif args.interactive:
        print()
        print(painter.dim(
            "  No --out given; manifest not persisted. Re-run with --out to save."
        ))

    sys.stdout.flush()
    # Exit code: 1 if any clusters > 1 found (split is constrained), 0 if clean.
    return 1 if manifest.clusters else 0


if __name__ == "__main__":
    raise SystemExit(main())
