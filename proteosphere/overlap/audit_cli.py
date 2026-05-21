"""Command-line interface for auditing an already-existing split.

``python -m proteosphere audit-split --input split.csv``

Reads a list of (accession, split) pairs and reports any leakage cluster
whose members span more than one split. This is the post-hoc complement
to ``overlap-cluster``: instead of *building* constraints, it *checks*
whether an existing split honours them.

The audit is conceptually simple:

1. Extract every accession + its split label from the input.
2. Run :func:`proteosphere.overlap.compute_leakage_clusters` over the
   union under the chosen tiers and specificity cap.
3. For each cluster, check which splits its members occupy. A cluster
   landing in >=2 splits is a violation; the cluster's source list
   tells you *which* axis caused the link.

Input formats
-------------

The auditor accepts:

* CSV / TSV / XLSX with ``accession`` (or alias) + ``split`` columns.
* JSON ``{"train": [...], "val": [...], "test": [...]}``.
* JSON ``{"records": [{"accession": "...", "split": "..."}, ...]}``.
* The freestanding-app DatasetManifest format (``records`` with
  ``protein_a`` / ``protein_b`` + ``split`` fields).
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from proteosphere.config import Config

from .clusters import (
    DEFAULT_CLUSTER_TIERS,
    DEFAULT_MIN_SPECIFICITY,
    LeakageCluster,
    LeakageManifest,
    TIER_SOURCES,
    compute_leakage_clusters,
)
from .cli import (
    _Ansi,
    _Painter,
    _enable_ansi_on_windows,
    _fmt_duration,
    _supports_color,
    _terminal_width,
    _truncate,
    _TIER_STYLE,
    resolve_config_or_friendly_exit,
    resolve_pdb_details,
)
from .cluster_cli import (
    _looks_like_id,
    _looks_like_header,
    _looks_like_pdb,
    _split_id_cell,
)
from .column_aliases import (
    ACCESSION_COLUMNS,
    COMPARISON_COLUMNS,
    PAIR_A_COLUMNS,
    PAIR_B_COLUMNS,
    PDB_COLUMNS,
    PROTEIN_COLUMNS,
    ROW_ID_COLUMNS,
    SPLIT_COLUMNS,
    TEST_COLUMNS,
    column_role,
    find_all_column_indices,
    find_column_index,
    normalize as _normalize_col,
)
from .tiers import SEVERITY_TIERS


# ---------------------------------------------------------------------------
# Input loading
# ---------------------------------------------------------------------------

# Roles the audit-split CLI cares about. We bind to the centralized
# aliases so a user's column named ``Test IDs`` / ``test_pdb_id`` /
# ``training set`` / ``CV Fold`` all match.
_SPLIT_ALIASES = SPLIT_COLUMNS
_ACC_ALIASES = (
    ACCESSION_COLUMNS | PDB_COLUMNS | PROTEIN_COLUMNS
    | ROW_ID_COLUMNS
)
# Columns that carry IDs but ALSO encode pair-side / split-side context
# (e.g. ``test_id`` and ``protein_a``). We sweep these as a fallback so
# pair manifests with split labels still parse.
_PAIR_LIKE_ALIASES = (
    TEST_COLUMNS | COMPARISON_COLUMNS
    | PAIR_A_COLUMNS | PAIR_B_COLUMNS
)


def _load_csv_split(path: Path) -> list[tuple[str, str]]:
    """Parse a CSV/TSV file with accession + split columns.

    Column-name detection is fuzzy: ``Split``, ``split``, ``Fold``,
    ``Partition``, ``CV Fold``, etc. all match the split column;
    ``test_id``, ``Test ID``, ``test_ids``, ``train_id``, ``comp_id``,
    ``protein_a``, ``protein_b``, ``accession``, ``pdb_id`` etc. all
    match accession-bearing columns.

    Two distinct input shapes are supported:

    1. **Explicit split column.** Most common: each row has an
       accession + a ``split`` label. A row's tokens all inherit that
       row's split value.
    2. **Implicit split via column name.** Paired benchmarks
       (``train_id``, ``test_id``; ``test_pdb_id``, ``comp_pdb_id``;
       etc.) have no split column - the split label is the column
       itself. The loader maps test-role columns to ``split="test"``
       and comparison-role columns to ``split="train"``.
    """
    delim = "\t" if path.suffix.lower() in (".tsv", ".txt") else ","
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh, delimiter=delim))
    if not rows:
        return []
    header = list(rows[0])
    if not _looks_like_header(header):
        raise SystemExit(
            f"{path}: first row does not look like a header. The auditor "
            f"needs an accession-bearing column and a split column. "
            f"Pass a JSON file if your data has no header row."
        )

    split_col = find_column_index(header, _SPLIT_ALIASES)
    acc_col = find_column_index(header, _ACC_ALIASES)
    # Any pair-side / test-side columns ALSO carry IDs; sweep them all
    # so files like (comp_id, test_id, split) or (protein_a, protein_b,
    # split) parse correctly.
    pair_cols = find_all_column_indices(header, _PAIR_LIKE_ALIASES)

    # Shape (2): no explicit split column but the test/comparison
    # columns themselves encode the split. Each ID-bearing column
    # contributes its role-derived split label.
    if split_col is None:
        inferred: list[tuple[int, str]] = []  # (col_index, split_label)
        for i, name in enumerate(header):
            role = column_role(name)
            if role == "test":
                inferred.append((i, "test"))
            elif role == "comparison":
                inferred.append((i, "train"))
        # Inferred split is only meaningful if we have at least two
        # distinct split labels (otherwise every row would be tagged
        # "test" and no audit can fire).
        inferred_labels = {label for _, label in inferred}
        if len(inferred_labels) >= 2:
            out: list[tuple[str, str]] = []
            for row in rows[1:]:
                if not row:
                    continue
                for col, label in inferred:
                    if col < len(row):
                        for tok in _split_id_cell(row[col]):
                            out.append((tok, label))
            return out
        raise SystemExit(
            f"{path}: could not find a split column.\n"
            f"  Acceptable split-column names include: split, splits, "
            f"fold, folds, partition, set, subset, assignment, "
            f"data_split, ml_split.\n"
            f"  Or, for paired-benchmark files with no split column, "
            f"use a test-side column AND a train/comparison-side column "
            f"(e.g. 'train_id' + 'test_id', 'comp_pdb_id' + 'test_pdb_id').\n"
            f"  Headers seen: {header!r}"
        )

    if acc_col is None and not pair_cols:
        raise SystemExit(
            f"{path}: could not find any accession-bearing column.\n"
            f"  Acceptable names include any of:\n"
            f"    UniProt:   accession, accessions, uniprot, uniprot_id\n"
            f"    PDB:       pdb, pdb_id, structure_id, complex_id\n"
            f"    Protein:   protein, protein_id, gene, gene_id\n"
            f"    Pair:      test_id, comp_id, train_id, protein_a, protein_b\n"
            f"    Row-id:    row_id, record_id, complex_id\n"
            f"  Headers seen: {header!r}"
        )

    out: list[tuple[str, str]] = []
    for row in rows[1:]:
        if not row:
            continue
        if split_col >= len(row):
            continue
        split_value = (row[split_col] or "").strip()
        if not split_value:
            continue
        if acc_col is not None and acc_col < len(row):
            for tok in _split_id_cell(row[acc_col]):
                out.append((tok, split_value))
        for col in pair_cols:
            if col < len(row):
                for tok in _split_id_cell(row[col]):
                    out.append((tok, split_value))
    return out


def _load_json_split(path: Path) -> list[tuple[str, str]]:
    """Parse a JSON file with any of the supported split-manifest shapes."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    out: list[tuple[str, str]] = []

    def _add(acc: object, split: str) -> None:
        if not acc:
            return
        for tok in _split_id_cell(str(acc)):
            out.append((tok, split))

    # Shape 1: top-level mapping of split-name -> [accessions]
    if isinstance(payload, dict) and all(
        isinstance(v, list) for v in payload.values()
    ) and any(k in {"train", "val", "test", "fold_0", "fold_1"}
              for k in payload):
        for split_name, accs in payload.items():
            for a in accs:
                _add(a, str(split_name))
        return out

    # Shape 2: object with "records" (DatasetManifest style)
    if isinstance(payload, dict) and "records" in payload:
        records = payload["records"]
    elif isinstance(payload, list):
        records = payload
    else:
        raise SystemExit(
            f"{path}: unrecognised JSON shape. Expected a top-level "
            f"object with 'records', a bare list of records, or a "
            f"mapping of split-name -> accession list."
        )

    for r in records:
        if not isinstance(r, dict):
            continue
        split = r.get("split") or r.get("fold") or r.get("partition")
        if not split:
            continue
        # Pull every accession-shaped field we recognise.
        for key in ("accession", "protein_a", "protein_b", "test_id",
                    "comparison_id", "a", "b"):
            val = r.get(key)
            if isinstance(val, list):
                for v in val:
                    _add(v, str(split))
            elif val:
                _add(val, str(split))
    return out


def _load_xlsx_split(path: Path) -> list[tuple[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install with:\n"
            "  pip install openpyxl"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    # Reuse the CSV path via a tempfile-like translation.
    text = "\n".join(
        ",".join("" if c is None else str(c).replace(",", " ") for c in row)
        for row in rows
    )
    import io
    rows_csv = list(csv.reader(io.StringIO(text)))
    header = [h.strip().lower() for h in rows_csv[0]]
    if not _looks_like_header(header):
        raise SystemExit(f"{path}: first sheet row does not look like a header.")
    # Reuse the same logic by writing a temporary CSV in memory.
    # (Simpler than duplicating; just delegate.)
    import tempfile
    with tempfile.NamedTemporaryFile(
        "w", suffix=".csv", delete=False, encoding="utf-8",
    ) as tmp:
        tmp_path = Path(tmp.name)
        tmp.write(text)
    try:
        return _load_csv_split(tmp_path)
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


def load_split(path: Path) -> list[tuple[str, str]]:
    """Dispatch by file extension; returns (accession, split) tuples."""
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        return _load_csv_split(path)
    if suffix == ".json":
        return _load_json_split(path)
    if suffix == ".xlsx":
        return _load_xlsx_split(path)
    raise SystemExit(
        f"Unsupported input format: {suffix!r}. Supported: .csv .tsv .json .xlsx"
    )


# ---------------------------------------------------------------------------
# Auditing
# ---------------------------------------------------------------------------


@dataclass
class _Violation:
    """A leakage cluster that landed in more than one split."""
    cluster: LeakageCluster
    members_by_split: dict[str, list[str]]


def find_violations(
    manifest: LeakageManifest,
    accession_to_split: dict[str, str],
) -> list[_Violation]:
    """Return one entry per cluster that crosses split boundaries."""
    out: list[_Violation] = []
    for cluster in manifest.clusters:
        members_by_split: dict[str, list[str]] = defaultdict(list)
        for acc in cluster.members:
            split = accession_to_split.get(acc)
            if split is None:
                continue
            members_by_split[split].append(acc)
        if len(members_by_split) >= 2:
            out.append(_Violation(
                cluster=cluster,
                members_by_split=dict(sorted(members_by_split.items())),
            ))
    # Sort largest violations first.
    out.sort(
        key=lambda v: (-sum(len(m) for m in v.members_by_split.values()),
                       v.cluster.cluster_id),
    )
    return out


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------


def _render_header(
    *, painter: _Painter, input_path: Path, n_records: int,
    splits: dict[str, int], tiers: list[str], min_specificity: int | None,
    warehouse_root: Path, width: int,
) -> None:
    title = "ProteoSphere Split Auditor"
    print(painter.bold(title))
    print("=" * len(title))
    print()
    split_summary = ", ".join(f"{name}={count}" for name, count in sorted(splits.items()))
    fields = [
        ("Input          ", f"{input_path}  ({n_records} records)"),
        ("Splits         ", split_summary),
        ("Tiers          ", ", ".join(tiers)),
        ("Min specificity",
         f"{min_specificity}" if min_specificity is not None else "no cap"),
        ("Warehouse      ", str(warehouse_root)),
    ]
    for name, value in fields:
        line = f"  {name}  {value}"
        if len(line) > width:
            line = line[: width - 3] + "..."
        print(line)
    print()


def _render_violations(
    *, painter: _Painter, violations: list[_Violation], max_show: int, width: int,
) -> None:
    if not violations:
        print(painter.bold("No cross-split leakage detected at the chosen tiers."))
        return
    print(painter.bold(
        f"Found {len(violations)} cross-split leakage cluster(s):"
    ))
    print("-" * 78)
    for v in violations[:max_show]:
        cluster = v.cluster
        # Severity tier of the strongest source.
        top_source = cluster.sources[0] if cluster.sources else None
        top_tier = top_source.tier if top_source else "?"
        # Render split distribution as e.g. "train=2, test=1".
        dist = ", ".join(
            f"{painter.tier(top_tier, k)}={len(vlist)}"
            for k, vlist in v.members_by_split.items()
        )
        print(f"  {cluster.cluster_id}  {painter.tier(top_tier, top_tier):<22}  "
              f"splits={dist}")
        # Per-split members
        for split_name, members in v.members_by_split.items():
            sample = ", ".join(members[:6])
            extra = "" if len(members) <= 6 else f", +{len(members)-6} more"
            line = f"      [{split_name:<8}] {sample}{extra}"
            if len(line) > width:
                line = line[: width - 3] + "..."
            print(painter.dim(line))
        # Top source identifier
        if top_source:
            via = (
                f"      via {top_source.tier}/{top_source.namespace}:"
                f"{top_source.identifier}"
            )
            if top_source.prevalence:
                via += f" [worldwide={top_source.prevalence}]"
            if len(cluster.sources) > 1:
                via += f"  (+{len(cluster.sources)-1} more source(s))"
            print(painter.dim(via))
        print()
    if len(violations) > max_show:
        print(painter.dim(
            f"  ... {len(violations) - max_show} more violation(s); see --out JSON for the full list."
        ))


def _render_summary(
    *, painter: _Painter, violations: list[_Violation],
    splits: dict[str, int],
) -> None:
    print()
    print(painter.bold("Summary"))
    print("-------")
    if not violations:
        print(painter.bold("  VERDICT: clean."))
        print("  No leakage cluster spans more than one split.")
        return
    # Tier counts
    tier_counts: dict[str, int] = defaultdict(int)
    n_records_affected = 0
    for v in violations:
        if v.cluster.sources:
            tier_counts[v.cluster.sources[0].tier] += 1
        else:
            tier_counts["unknown"] += 1
        n_records_affected += sum(len(m) for m in v.members_by_split.values())
    print(painter.bold(f"  VERDICT: leakage detected ({len(violations)} clusters)."))
    print(f"  Records involved in cross-split leakage: {n_records_affected}")
    print()
    print("  Violations by severity tier:")
    for tier in SEVERITY_TIERS:
        n = tier_counts.get(tier, 0)
        if n == 0:
            continue
        _color, marker = _TIER_STYLE.get(tier, (_Ansi.DIM, "   "))
        line = f"    {tier:<28} {n:>4}   {marker}"
        print(painter.tier(tier, line))


# ---------------------------------------------------------------------------
# Argparse + main
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="proteosphere audit-split",
        description=(
            "Audit an already-existing train/val/test (or k-fold) split for "
            "biological-similarity leakage. Reads (accession, split) tuples, "
            "builds leakage clusters under the chosen tiers, and reports "
            "every cluster whose members span more than one split."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  proteosphere audit-split --input split.csv\n"
            "  proteosphere audit-split --input split.json --tiers identity,direct_ortholog,paralog_family,convergent_function\n"
            "  proteosphere audit-split --input split.csv --out audit.json\n"
        ),
    )
    parser.add_argument(
        "--input", "-i", type=Path,
        help="CSV / TSV / JSON / XLSX with (accession, split) rows.",
    )
    parser.add_argument(
        "--tiers", default=",".join(DEFAULT_CLUSTER_TIERS),
        help=(
            "Comma-separated tiers to audit against. Default: "
            f"{','.join(DEFAULT_CLUSTER_TIERS)}"
        ),
    )
    parser.add_argument(
        "--min-specificity", type=int, default=DEFAULT_MIN_SPECIFICITY,
        help=(
            "Drop clustering identifiers whose worldwide member count exceeds "
            f"this cap (default: {DEFAULT_MIN_SPECIFICITY}). Pass 0 to disable."
        ),
    )
    parser.add_argument(
        "--out", "-o", type=Path, default=None,
        help="Write a JSON audit report (one entry per violation) to this path.",
    )
    parser.add_argument(
        "--max-show", type=int, default=20,
        help="Maximum number of violations to display on screen (default 20).",
    )
    parser.add_argument(
        "--id-kind", default="auto", choices=("auto", "accession", "pdb"),
        help=(
            "Type of identifier in the input. 'accession' (default for "
            "non-PDB-looking inputs) treats them as UniProt accessions; "
            "'pdb' resolves PDB IDs to UniProt accessions via the "
            "structure_units table before clustering. 'auto' (default) "
            "promotes to 'pdb' when the majority of input IDs look like "
            "4-character PDB codes."
        ),
    )
    parser.add_argument(
        "--no-color", action="store_true",
        help="Disable ANSI color output.",
    )
    return parser


def _resolve_tiers(text: str) -> list[str]:
    tiers = [t.strip() for t in text.split(",") if t.strip()]
    valid = set(TIER_SOURCES)
    unknown = [t for t in tiers if t not in valid]
    if unknown:
        raise SystemExit(
            f"Unknown tier(s): {unknown}\nValid tiers: {sorted(valid)}"
        )
    return tiers


def main(argv: list[str] | None = None, *, config: Config | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    use_color = not args.no_color and _supports_color()
    if use_color:
        _enable_ansi_on_windows()
    painter = _Painter(enabled=use_color)

    if args.input is None:
        parser.error("--input is required.")

    input_path: Path = args.input.expanduser().resolve()
    if not input_path.is_file():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        return 2

    try:
        rows = load_split(input_path)
    except (ValueError, json.JSONDecodeError) as e:
        print(f"ERROR: failed to read {input_path}:\n  {e}", file=sys.stderr)
        return 2
    if not rows:
        print(f"WARNING: no (accession, split) rows found in {input_path}.",
              file=sys.stderr)
        return 0

    # ----- Resolve input ID kind (accession vs PDB) -----
    # The cluster engine is keyed on UniProt accessions. If the user
    # handed us PDB codes (or a paired benchmark like splittest.xlsx
    # with PDB-formatted train_id / test_id columns), every token
    # would otherwise become a singleton and the audit would falsely
    # report "no leakage."  Resolve PDBs through structure_units
    # before clustering; each PDB's split label transfers to every
    # UniProt accession that PDB resolves to.
    id_kind = args.id_kind
    unique_tokens_in_order: list[str] = []
    seen_tokens: set[str] = set()
    for tok, _ in rows:
        t = (tok or "").strip().upper()
        if not t or t in seen_tokens:
            continue
        seen_tokens.add(t)
        unique_tokens_in_order.append(t)

    if id_kind == "auto":
        sample = unique_tokens_in_order[:200]
        n_pdb = sum(1 for t in sample if _looks_like_pdb(t))
        pdb_fraction = n_pdb / len(sample) if sample else 0.0
        if pdb_fraction >= 0.5:
            id_kind = "pdb"
            print(painter.dim(
                f"  Input looks PDB-dominant "
                f"({n_pdb}/{len(sample)} sampled IDs are 4-char PDB codes). "
                f"Resolving to UniProt via structure_units."
            ))
        else:
            id_kind = "accession"

    if id_kind == "pdb":
        # Resolve config now so we can use the warehouse for resolution.
        config = resolve_config_or_friendly_exit(config, painter=painter)
        details = resolve_pdb_details(config, unique_tokens_in_order)
        pdb_to_uniprots: dict[str, list[str]] = {
            d.requested_pdb.upper(): list(d.uniprots) for d in details
        }
        unresolved = [d.requested_pdb for d in details if not d.resolved]
        if unresolved:
            print(painter.dim(
                f"  warning: {len(unresolved)} PDB ID(s) did not resolve to "
                f"any UniProt accession; they are excluded from the audit."
            ))
            sample_un = ", ".join(unresolved[:6])
            extra = "" if len(unresolved) <= 6 else f", +{len(unresolved)-6} more"
            print(painter.dim(f"    {sample_un}{extra}"))
        # Expand each (PDB, split) into one row per resolved UniProt.
        expanded: list[tuple[str, str]] = []
        for tok, split in rows:
            t = (tok or "").strip().upper()
            for acc in pdb_to_uniprots.get(t, []):
                expanded.append((acc, split))
        if not expanded:
            print(
                "ERROR: no UniProt accessions could be resolved from the "
                "input PDB IDs. Pass --id-kind accession if the file "
                "actually contains UniProt accessions.",
                file=sys.stderr,
            )
            return 2
        rows = expanded

    # Index accessions by the split they appear in. An accession that
    # appears in more than one split row is itself a hard leakage at
    # the identity tier and gets flagged separately.
    accession_to_splits: dict[str, set[str]] = defaultdict(set)
    accessions: list[str] = []
    seen: set[str] = set()
    for acc, split in rows:
        acc_norm = acc.strip().upper()
        if not acc_norm:
            continue
        accession_to_splits[acc_norm].add(split.strip())
        if acc_norm not in seen:
            seen.add(acc_norm)
            accessions.append(acc_norm)

    # Identity-level violation: the same accession is in >=2 splits.
    identity_crossing = sorted(
        a for a, s in accession_to_splits.items() if len(s) >= 2
    )

    # For the cluster-based audit, take the "first" split per accession;
    # identity-crossings are reported separately.
    accession_to_split: dict[str, str] = {
        a: next(iter(s)) for a, s in accession_to_splits.items()
    }

    split_counts = defaultdict(int)
    for s in accession_to_split.values():
        split_counts[s] += 1

    config = resolve_config_or_friendly_exit(config, painter=painter)

    tiers = _resolve_tiers(args.tiers)
    min_specificity = None if args.min_specificity <= 0 else args.min_specificity
    width = _terminal_width()

    _render_header(
        painter=painter, input_path=input_path,
        n_records=len(accessions), splits=dict(split_counts),
        tiers=tiers, min_specificity=min_specificity,
        warehouse_root=config.warehouse_root, width=width,
    )

    t0 = time.monotonic()
    print(painter.dim(
        "Computing leakage clusters over the union... "
        "(first run pays the cold-cache cost)"
    ))
    cluster_manifest = compute_leakage_clusters(
        config, accessions, tiers=tiers, min_specificity=min_specificity,
    )
    elapsed = time.monotonic() - t0
    print(painter.dim(f"  computed in {_fmt_duration(elapsed)}"))
    print()

    violations = find_violations(cluster_manifest, accession_to_split)

    # Identity crossings get their own pseudo-violation block.
    if identity_crossing:
        print(painter.tier("identity", painter.bold(
            f"IDENTITY-LEVEL LEAKAGE: {len(identity_crossing)} accession(s) "
            f"appear in more than one split"
        )))
        for acc in identity_crossing[:10]:
            splits = sorted(accession_to_splits[acc])
            print(painter.tier(
                "identity",
                f"  {acc}  splits={splits}",
            ))
        if len(identity_crossing) > 10:
            print(painter.dim(
                f"  ... and {len(identity_crossing) - 10} more."
            ))
        print()

    _render_violations(
        painter=painter, violations=violations,
        max_show=args.max_show, width=width,
    )
    _render_summary(
        painter=painter, violations=violations, splits=dict(split_counts),
    )

    if args.out is not None:
        out_path: Path = args.out.expanduser().resolve()
        payload = {
            "input": str(input_path),
            "tiers": tiers,
            "min_specificity": min_specificity,
            "split_counts": dict(split_counts),
            "identity_crossings": [
                {"accession": acc, "splits": sorted(accession_to_splits[acc])}
                for acc in identity_crossing
            ],
            "cluster_violations": [
                {
                    "cluster_id": v.cluster.cluster_id,
                    "size": len(v.cluster.members),
                    "members_by_split": v.members_by_split,
                    "sources": [
                        {
                            "tier": s.tier, "namespace": s.namespace,
                            "identifier": s.identifier,
                            "worldwide_members": s.prevalence,
                        }
                        for s in v.cluster.sources
                    ],
                }
                for v in violations
            ],
        }
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        print()
        print(f"  Wrote audit report -> {out_path}")

    sys.stdout.flush()
    # Exit 1 if any violation found, 0 if clean.
    return 1 if (violations or identity_crossing) else 0


if __name__ == "__main__":
    raise SystemExit(main())
