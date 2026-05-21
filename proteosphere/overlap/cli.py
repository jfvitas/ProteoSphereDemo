"""Command-line interface for the overlap scanner.

Usage::

    python -m proteosphere overlap-scan --input pairs.csv \\
        --profile structure_prediction --out report.json

The CLI reads a list of pairs from a CSV / TSV / JSON / XLSX file, runs
each through the multi-axis overlap engine under the chosen task profile,
and prints a sorted, color-coded summary to the terminal. A full JSON
report (one entry per pair) can be written with ``--out``.

Supported input formats
-----------------------

CSV / TSV (header row required):

    test_id,comparison_id,label
    P00639,P60709,actin_pair
    P00760,P00782,trypsin_vs_subtilisin

Multiple accessions per cell can be joined with ``;`` or ``|`` or ``,``.
The ``label`` column is optional. PDB IDs work too with ``--id-kind pdb``;
the scanner resolves them to UniProt via the ``structure_units`` table.

JSON:

    {
      "pairs": [
        {"label": "actin", "test": ["P00639", "P68135"], "comparison": ["P60709"]},
        ...
      ]
    }

A bare list of pair objects is also accepted.

XLSX: same columns as CSV, first sheet only. Requires the optional
``openpyxl`` dependency (``pip install proteosphere[xlsx]``).
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sys
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence

from proteosphere.config import Config

from .profiles import TASK_PROFILES, get_task_profile
from .report import OverlapReport
from .runner import discover_overlap
from .tiers import SEVERITY_TIERS, TIER_DESCRIPTIONS


# ---------------------------------------------------------------------------
# Defaults & layout constants
# ---------------------------------------------------------------------------

DEFAULT_THRESHOLD = 0.10
DEFAULT_PROFILE = "all"
DEFAULT_TERM_WIDTH = 100
MIN_TERM_WIDTH = 80

# Cell separator for multi-accession columns in CSV/XLSX.
ID_SEPARATORS = (";", "|", ",")


# ---------------------------------------------------------------------------
# Colors (ANSI). Used only when stdout is a TTY and the user hasn't passed
# --no-color. We pick from the basic 8-color palette so legacy cmd.exe (with
# Windows 10+ VT processing enabled) renders correctly.
# ---------------------------------------------------------------------------

class _Ansi:
    RESET = "\x1b[0m"
    BOLD = "\x1b[1m"
    DIM = "\x1b[2m"
    RED = "\x1b[31m"
    GREEN = "\x1b[32m"
    YELLOW = "\x1b[33m"
    BLUE = "\x1b[34m"
    MAGENTA = "\x1b[35m"
    CYAN = "\x1b[36m"
    GRAY = "\x1b[90m"
    BRIGHT_RED = "\x1b[91m"
    BRIGHT_YELLOW = "\x1b[93m"


# Tier -> (label color, severity weight). Severity weight is informational
# only — used for the *** / ** / * markers in the summary.
_TIER_STYLE: dict[str, tuple[str, str]] = {
    "identity":                    (_Ansi.BRIGHT_RED + _Ansi.BOLD, "***"),
    "direct_ortholog":             (_Ansi.BRIGHT_RED,              "***"),
    "paralog_family":              (_Ansi.YELLOW,                  "** "),
    "domain_architecture":         (_Ansi.YELLOW,                  "** "),
    "distant_homology":            (_Ansi.BRIGHT_YELLOW,           "** "),
    "convergent_function":         (_Ansi.CYAN,                    "** "),
    "broad_fold":                  (_Ansi.BLUE,                    "*  "),
    "shared_motif":                (_Ansi.BLUE,                    "*  "),
    "shared_partial_architecture": (_Ansi.GRAY,                    "*  "),
    "co_localization":             (_Ansi.GRAY,                    "*  "),
    "shared_pathway":              (_Ansi.GRAY,                    "*  "),
    "shared_partner":              (_Ansi.GRAY,                    "*  "),
    "none":                        (_Ansi.DIM,                     "   "),
    "unknown":                     (_Ansi.DIM,                     "   "),
}


# ---------------------------------------------------------------------------
# Pair input model
# ---------------------------------------------------------------------------


@dataclass
class PairSpec:
    """One pair to scan: optional label + test/comparison ID lists."""
    label: str
    test_ids: list[str]
    comparison_ids: list[str]
    row_number: int = 0  # 1-based index in source file, for diagnostics


# ---------------------------------------------------------------------------
# File loaders
# ---------------------------------------------------------------------------


# Use the centralized alias registry so the three CLIs share the same
# vocabulary. Roles used here:
#   - TEST_COLUMNS: the held-out / evaluation side of a pair
#   - COMPARISON_COLUMNS: the train / reference side of a pair
#   - LABEL_COLUMNS: free-text label / pair name
# Plus PAIR_A_COLUMNS / PAIR_B_COLUMNS for symmetric (`protein_a` /
# `protein_b`) pair shapes.
from .column_aliases import (
    COMPARISON_COLUMNS as _COMP_ALIASES,
    LABEL_COLUMNS as _LABEL_ALIASES,
    PAIR_A_COLUMNS as _PAIR_A_ALIASES,
    PAIR_B_COLUMNS as _PAIR_B_ALIASES,
    TEST_COLUMNS as _TEST_ALIASES,
    normalize as _normalize_col,
)


def _split_ids(cell: str) -> list[str]:
    """Split a CSV cell into individual IDs, accepting ``;`` ``|`` or ``,``."""
    if cell is None:
        return []
    cell = str(cell).strip()
    if not cell:
        return []
    # Try the most-specific separators first; fall back to comma if neither
    # ``;`` nor ``|`` is present (so ``PXXXX,PYYYY`` works inside one cell).
    for sep in ID_SEPARATORS:
        if sep in cell:
            return [t.strip() for t in cell.split(sep) if t.strip()]
    return [cell]


def _normalize_header(name: str) -> str:
    """Compatibility shim -- delegate to the centralized normalizer."""
    return _normalize_col(name)


def _resolve_columns(headers: Sequence[str]) -> tuple[int, int, int | None]:
    """Map a row of headers to (test_col, comp_col, label_col_or_None).

    Tries, in order:

    1. A column matching a test-role alias (e.g. ``test_id``, ``test_ids``,
       ``query``) paired with a comparison-role alias (``comp_id``,
       ``train_id``, ``reference``).
    2. The symmetric pair shape ``protein_a`` / ``protein_b`` (or
       ``side_a`` / ``side_b``, ``first`` / ``second``, etc.) — here
       there's no test/comparison semantics, so we treat ``a`` as
       test and ``b`` as comparison.

    Raises ValueError with an actionable message if neither shape is
    recognised.
    """
    norm = [_normalize_col(h) for h in headers]
    test_col = next((i for i, h in enumerate(norm) if h in _TEST_ALIASES), -1)
    comp_col = next((i for i, h in enumerate(norm) if h in _COMP_ALIASES), -1)

    # Fallback to symmetric pair shape if explicit test/comp missing.
    if test_col == -1 or comp_col == -1:
        a_col = next((i for i, h in enumerate(norm) if h in _PAIR_A_ALIASES), -1)
        b_col = next((i for i, h in enumerate(norm) if h in _PAIR_B_ALIASES), -1)
        if a_col != -1 and b_col != -1:
            test_col, comp_col = a_col, b_col

    label_col = next((i for i, h in enumerate(norm) if h in _LABEL_ALIASES), None)
    if test_col == -1 or comp_col == -1:
        # Build a friendlier error showing a representative sample of
        # aliases the user can rename their column to.
        sample_test = sorted(_TEST_ALIASES)[:6]
        sample_comp = sorted(_COMP_ALIASES)[:6]
        raise ValueError(
            f"Could not find required columns. Headers seen: {list(headers)!r}.\n"
            f"Need a 'test' column (any of: {sample_test} ...) AND a "
            f"'comparison' column (any of: {sample_comp} ...), OR a "
            f"symmetric pair shape with 'protein_a' + 'protein_b' / "
            f"'first' + 'second' / 'side_a' + 'side_b'."
        )
    return test_col, comp_col, label_col


def _load_csv(path: Path, *, delim: str = ",") -> list[PairSpec]:
    pairs: list[PairSpec] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=delim)
        rows = list(reader)
    if not rows:
        return pairs
    test_col, comp_col, label_col = _resolve_columns(rows[0])
    for i, row in enumerate(rows[1:], start=2):  # 2 = first data row, 1-based
        if not row or all(not (c or "").strip() for c in row):
            continue
        # Pad short rows so we don't IndexError on missing trailing cells.
        if len(row) <= max(test_col, comp_col, label_col or 0):
            row = list(row) + [""] * (max(test_col, comp_col, label_col or 0) + 1 - len(row))
        test_ids = _split_ids(row[test_col])
        comp_ids = _split_ids(row[comp_col])
        if not test_ids or not comp_ids:
            continue
        label = (row[label_col].strip() if label_col is not None else "")
        pairs.append(PairSpec(
            label=label or f"row_{i}",
            test_ids=test_ids,
            comparison_ids=comp_ids,
            row_number=i,
        ))
    return pairs


def _load_tsv(path: Path) -> list[PairSpec]:
    return _load_csv(path, delim="\t")


def _load_json(path: Path) -> list[PairSpec]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and "pairs" in payload:
        items = payload["pairs"]
    elif isinstance(payload, list):
        items = payload
    else:
        raise ValueError(
            "JSON input must be a list of pair objects, or an object with a "
            "'pairs' list."
        )
    pairs: list[PairSpec] = []
    for i, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"JSON pair #{i} is not an object: {item!r}")
        test_raw = item.get("test") or item.get("test_id") or item.get("a")
        comp_raw = (
            item.get("comparison") or item.get("comparison_id")
            or item.get("comp") or item.get("b")
        )
        if test_raw is None or comp_raw is None:
            raise ValueError(
                f"JSON pair #{i} missing 'test' and/or 'comparison' field: "
                f"{item!r}"
            )
        test_ids = _coerce_id_list(test_raw)
        comp_ids = _coerce_id_list(comp_raw)
        if not test_ids or not comp_ids:
            continue
        label = str(item.get("label") or item.get("name") or f"pair_{i}")
        pairs.append(PairSpec(
            label=label,
            test_ids=test_ids,
            comparison_ids=comp_ids,
            row_number=i,
        ))
    return pairs


def _coerce_id_list(value: object) -> list[str]:
    """Accept a list or a delimited string; return a normalized list of IDs."""
    if isinstance(value, list):
        out: list[str] = []
        for v in value:
            if v is None:
                continue
            out.extend(_split_ids(str(v)))
        return out
    if isinstance(value, str):
        return _split_ids(value)
    return _split_ids(str(value))


def _load_xlsx(path: Path) -> list[PairSpec]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install with:\n"
            "  pip install openpyxl\n"
            "or:  pip install proteosphere[xlsx]"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    test_col, comp_col, label_col = _resolve_columns(headers)
    pairs: list[PairSpec] = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # Same padding logic as CSV.
        if len(row) <= max(test_col, comp_col, label_col or 0):
            row = list(row) + [""] * (max(test_col, comp_col, label_col or 0) + 1 - len(row))
        test_ids = _split_ids("" if row[test_col] is None else str(row[test_col]))
        comp_ids = _split_ids("" if row[comp_col] is None else str(row[comp_col]))
        if not test_ids or not comp_ids:
            continue
        if label_col is not None and row[label_col] is not None:
            label = str(row[label_col]).strip()
        else:
            label = ""
        pairs.append(PairSpec(
            label=label or f"row_{i}",
            test_ids=test_ids,
            comparison_ids=comp_ids,
            row_number=i,
        ))
    return pairs


def load_pairs(path: Path) -> list[PairSpec]:
    """Dispatch on file extension and load a list of pairs."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in (".tsv", ".txt"):
        return _load_tsv(path)
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    raise SystemExit(
        f"Unsupported input format: {suffix!r}. Supported: .csv .tsv .json .xlsx"
    )


# ---------------------------------------------------------------------------
# PDB -> UniProt resolution (when --id-kind pdb)
# ---------------------------------------------------------------------------


def _resolve_pdb_set(config: Config, pdb_ids: Iterable[str]) -> dict[str, list[str]]:
    """Resolve PDB IDs to their UniProt accessions, supplement-aware.

    Returns ``{requested_pdb_upper: [accession, ...]}``. Missing PDBs
    map to an empty list. Used by callers that don't need the per-PDB
    provenance metadata; callers that DO need it (e.g. the cluster CLI
    that wants to surface obsolete redirects and unresolvable reasons)
    should call :func:`resolve_pdb_details` instead.

    Resolution order, per PDB:

    1. Apply any obsolete-redirect (e.g. ``1A2K -> 5BXQ``) so legacy
       SKEMPI/SIFTS IDs continue to resolve.
    2. Look up the effective PDB in the catalog's ``structure_units``
       table.
    3. Merge in any supplementary mappings from
       ``proteosphere/data/pdb_resolution_supplement.tsv`` (recovers
       TrEMBL-only entries SIFTS hasn't indexed).
    """
    details = resolve_pdb_details(config, pdb_ids)
    return {d.requested_pdb: list(d.uniprots) for d in details}


def resolve_pdb_details(
    config: Config, pdb_ids: Iterable[str],
) -> list:
    """Resolve a batch of PDB IDs with full provenance metadata.

    Returns a list of :class:`pdb_supplement.ResolutionDetail` objects --
    one per input PDB, in input order (deduplicated, uppercased).
    Carries the effective PDB after redirect, the list of UniProt
    accessions, the sources that contributed them, and any known-
    unresolvable reason. This is what the CLI uses to print "we couldn't
    resolve PDB X because it's a synthetic Fab" messages instead of
    silently dropping the entry.
    """
    import duckdb
    from .pdb_supplement import apply_obsolete_redirect, merge_resolutions

    # Normalize, dedupe, and resolve any obsolete redirects up front so
    # we batch-query the catalog by *effective* PDB ID.
    seen_in_order: list[str] = []
    seen_set: set[str] = set()
    for p in pdb_ids:
        s = (p or "").strip().upper()
        if not s or s in seen_set:
            continue
        seen_set.add(s)
        seen_in_order.append(s)
    if not seen_in_order:
        return []

    effective_pdbs: set[str] = set()
    for p in seen_in_order:
        eff, _ = apply_obsolete_redirect(p)
        effective_pdbs.add(eff)

    sifts_map: dict[str, list[str]] = {p: [] for p in effective_pdbs}
    con = duckdb.connect(str(config.catalog_path()), read_only=True)
    try:
        ids_clause = ", ".join(
            "'" + p.replace("'", "''") + "'" for p in sorted(effective_pdbs)
        )
        rows = con.execute(
            f"SELECT structure_id, protein_ref FROM main.structure_units "
            f"WHERE UPPER(structure_id) IN ({ids_clause}) "
            f"AND protein_ref LIKE 'protein:%'"
        ).fetchall()
    finally:
        con.close()

    seen_acc: dict[str, set[str]] = {p: set() for p in effective_pdbs}
    for sid, protein_ref in rows:
        if not sid or not protein_ref:
            continue
        s = sid.upper()
        if s not in seen_acc:
            continue
        accession = protein_ref.replace("protein:", "")
        if accession and accession not in seen_acc[s]:
            seen_acc[s].add(accession)
            sifts_map[s].append(accession)
    for p in sifts_map:
        sifts_map[p].sort()
    return merge_resolutions(seen_in_order, sifts_map)


# ---------------------------------------------------------------------------
# Terminal helpers
# ---------------------------------------------------------------------------


def _enable_ansi_on_windows() -> None:
    """Best-effort: enable VT processing in legacy cmd.exe on Windows 10+."""
    if sys.platform != "win32":
        return
    try:  # pragma: no cover - Windows-specific
        import ctypes
        kernel32 = ctypes.windll.kernel32
        STD_OUTPUT_HANDLE = -11
        ENABLE_VIRTUAL_TERMINAL_PROCESSING = 0x0004
        handle = kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(handle, ctypes.byref(mode)):
            kernel32.SetConsoleMode(
                handle, mode.value | ENABLE_VIRTUAL_TERMINAL_PROCESSING
            )
    except Exception:
        pass


def _supports_color() -> bool:
    """True if stdout looks like a TTY that can render ANSI escapes."""
    if not sys.stdout.isatty():
        return False
    if os.environ.get("NO_COLOR"):
        return False
    return True


def _terminal_width() -> int:
    try:
        return max(MIN_TERM_WIDTH, shutil.get_terminal_size((DEFAULT_TERM_WIDTH, 24)).columns)
    except Exception:
        return DEFAULT_TERM_WIDTH


def _truncate(text: str, max_len: int) -> str:
    if max_len <= 0:
        return ""
    if len(text) <= max_len:
        return text
    if max_len <= 3:
        return text[:max_len]
    return text[: max_len - 3] + "..."


@dataclass
class _Painter:
    """Wraps text in ANSI escape codes when color is enabled."""
    enabled: bool

    def paint(self, text: str, color: str) -> str:
        if not self.enabled or not color:
            return text
        return f"{color}{text}{_Ansi.RESET}"

    def tier(self, tier: str, text: str | None = None) -> str:
        color, _ = _TIER_STYLE.get(tier, (_Ansi.DIM, "   "))
        return self.paint(text if text is not None else tier, color)

    def dim(self, text: str) -> str:
        return self.paint(text, _Ansi.DIM)

    def bold(self, text: str) -> str:
        return self.paint(text, _Ansi.BOLD)


def resolve_config_or_friendly_exit(
    config: Config | None,
    *,
    painter: _Painter,
) -> Config:
    """Return a usable Config, or print a friendly error and exit cleanly.

    The CLI subcommands call this instead of ``Config.discover()`` so the
    user never sees a raw Python traceback for the common "no warehouse
    configured" case. The message lists every place we looked and tells
    them exactly what to set.
    """
    if config is not None:
        return config
    try:
        return Config.discover()
    except FileNotFoundError as e:
        print()
        print(painter.bold("ERROR: ProteoSphere reference warehouse not found."))
        print()
        print("The scanner needs a path to the warehouse. We checked:")
        print("  - PROTEOSPHERE_WAREHOUSE environment variable")
        print("  - proteosphere_config.json in this folder or any parent")
        print("  - C:\\Users\\<you>\\.proteosphere\\reference_library  (default)")
        print()
        print("Fix it one of three ways:")
        print()
        print(painter.bold("  Option 1 (one-off): pass --warehouse-root"))
        print('    python -m proteosphere --warehouse-root "D:\\path\\to\\reference_library" \\')
        print('        overlap-scan ...')
        print()
        print(painter.bold("  Option 2 (this shell): set the env var"))
        print('    set PROTEOSPHERE_WAREHOUSE=D:\\path\\to\\reference_library')
        print('    (PowerShell: $env:PROTEOSPHERE_WAREHOUSE = "D:\\path\\to\\reference_library")')
        print()
        print(painter.bold("  Option 3 (permanent): set it for your user account"))
        print('    setx PROTEOSPHERE_WAREHOUSE "D:\\path\\to\\reference_library"')
        print('    (open a new shell after running setx)')
        print()
        print(painter.dim(f"  Detail: {e}"))
        sys.stdout.flush()
        raise SystemExit(2) from None


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------


@dataclass
class _ScannedPair:
    spec: PairSpec
    test_resolved: list[str]
    comparison_resolved: list[str]
    report: OverlapReport | None
    skip_reason: str | None = None


def _render_header(
    *, painter: _Painter, input_path: Path, n_pairs: int, profile: str,
    threshold: float, warehouse_root: Path, id_kind: str, width: int,
) -> None:
    title = "ProteoSphere Overlap Scanner"
    print(painter.bold(title))
    print("=" * len(title))
    print()
    fields = [
        ("Input    ", f"{input_path}  ({n_pairs} pair{'s' if n_pairs != 1 else ''})"),
        ("Profile  ", profile),
        ("ID kind  ", id_kind),
        ("Threshold", f"{threshold:.2f}  (pairs at or above this score are flagged)"),
        ("Warehouse", str(warehouse_root)),
    ]
    for name, value in fields:
        line = f"  {name}  {value}"
        # Trim very long warehouse paths so the header fits.
        if len(line) > width:
            line = line[: width - 3] + "..."
        print(line)
    print()


def _format_id_list(ids: list[str], max_len: int) -> str:
    """Join IDs with commas, truncating with '+N more' if too long."""
    if not ids:
        return "-"
    joined = ",".join(ids)
    if len(joined) <= max_len:
        return joined
    # Show first ID + "+N more" to keep the cell scannable.
    suffix_template = ",+{n} more"
    head_budget = max_len - len(suffix_template.format(n=99))
    if head_budget < 4:
        return _truncate(joined, max_len)
    head = ids[0]
    if len(head) > head_budget:
        head = head[: head_budget - 3] + "..."
    return f"{head}{suffix_template.format(n=len(ids)-1)}"


def _render_pair_table(
    *, painter: _Painter, pairs: list[_ScannedPair], threshold: float,
    width: int, verbose: bool,
) -> None:
    # Sort: highest score first, then severity-tier ordering.
    tier_index = {t: i for i, t in enumerate(SEVERITY_TIERS)}

    def sort_key(p: _ScannedPair) -> tuple[float, int]:
        if p.report is None:
            return (-1.0, 999)
        tier = p.report.severity_tier()
        return (-p.report.composite_score(), tier_index.get(tier, 999))

    sorted_pairs = sorted(pairs, key=sort_key)

    n_col = max(2, len(str(len(pairs))))
    score_col = 6   # "0.000"
    tier_col = 22   # widest tier name fits comfortably
    # Distribute remaining space between label/test/comparison.
    fixed = 2 + n_col + 2 + score_col + 2 + tier_col + 2
    remaining = max(width - fixed - 2, 30)
    label_col = max(18, min(32, remaining // 3))
    test_col_w = max(12, (remaining - label_col) // 2)
    comp_col_w = max(12, remaining - label_col - test_col_w)

    # Header
    header = (
        f"  {'#':<{n_col}}  {'Score':<{score_col}}  {'Tier':<{tier_col}}  "
        f"{'Label':<{label_col}}  {'Test':<{test_col_w}}  {'Comparison':<{comp_col_w}}"
    )
    print(painter.bold(header))
    print("  " + "-" * (len(header) - 2))

    for idx, scanned in enumerate(sorted_pairs, start=1):
        spec = scanned.spec
        if scanned.report is None:
            score_text = "  -- "
            tier_text = scanned.skip_reason or "skipped"
            tier_painted = painter.dim(_truncate(tier_text, tier_col).ljust(tier_col))
            label = _truncate(spec.label, label_col).ljust(label_col)
            test = _truncate(_format_id_list(spec.test_ids, test_col_w), test_col_w).ljust(test_col_w)
            comp = _truncate(_format_id_list(spec.comparison_ids, comp_col_w), comp_col_w).ljust(comp_col_w)
            row = (
                f"  {str(idx):<{n_col}}  {score_text:<{score_col}}  "
                f"{tier_painted}  {label}  {test}  {comp}"
            )
            print(painter.dim(row))
            continue
        report = scanned.report
        score = report.composite_score()
        tier = report.severity_tier() if score > 0 or report.identity_accessions else "none"
        if score < threshold and tier != "identity":
            tier_display = "noise_floor" if tier == "none" else tier
        else:
            tier_display = tier

        score_text = f"{score:.3f}".rjust(score_col)
        tier_text = _truncate(tier_display, tier_col).ljust(tier_col)
        label = _truncate(spec.label, label_col).ljust(label_col)
        test = _truncate(
            _format_id_list(scanned.test_resolved or spec.test_ids, test_col_w),
            test_col_w,
        ).ljust(test_col_w)
        comp = _truncate(
            _format_id_list(scanned.comparison_resolved or spec.comparison_ids, comp_col_w),
            comp_col_w,
        ).ljust(comp_col_w)

        # Color: tier color for the tier cell, score colored too if flagged.
        tier_painted = painter.tier(tier, tier_text)
        score_painted = (
            painter.tier(tier, score_text)
            if score >= threshold or tier == "identity"
            else painter.dim(score_text)
        )
        row = (
            f"  {str(idx):<{n_col}}  {score_painted}  {tier_painted}  "
            f"{label}  {test}  {comp}"
        )
        if score < threshold and tier != "identity":
            print(painter.dim(row))
        else:
            print(row)

        if verbose and report.axes:
            # Per-axis breakdown indented under the row, ordered by weight.
            axes_sorted = sorted(
                report.axes.items(), key=lambda kv: -kv[1].weight,
            )
            for axis_name, axis in axes_sorted:
                line = (
                    f"      {axis_name:<36} "
                    f"weight={axis.weight:5.3f}  "
                    f"ns={axis.namespace:<10} "
                    f"shared={','.join(axis.shared_ids[:3])}"
                )
                if len(line) > width:
                    line = line[: width - 3] + "..."
                print(painter.dim(line))


def _render_summary(
    *, painter: _Painter, pairs: list[_ScannedPair], threshold: float,
) -> None:
    print()
    print(painter.bold("Summary"))
    print("-------")

    # Tier counts (in canonical order). Pairs below threshold collapse to
    # "noise_floor" so the summary tells you "real signals" at a glance.
    counts: dict[str, int] = {tier: 0 for tier in SEVERITY_TIERS}
    counts["noise_floor"] = 0
    counts["skipped"] = 0
    n_flagged = 0
    n_evaluated = 0
    for p in pairs:
        if p.report is None:
            counts["skipped"] += 1
            continue
        n_evaluated += 1
        score = p.report.composite_score()
        tier = p.report.severity_tier()
        flagged = score >= threshold or tier == "identity"
        if flagged:
            n_flagged += 1
            counts[tier] = counts.get(tier, 0) + 1
        else:
            counts["noise_floor"] += 1

    print(f"  {'Tier':<32}{'Count':>6}   Severity")
    print(f"  {'-' * 30:<32}{'-' * 5:>6}   {'-' * 20}")
    for tier in SEVERITY_TIERS:
        n = counts.get(tier, 0)
        if n == 0:
            continue
        _color, marker = _TIER_STYLE.get(tier, (_Ansi.DIM, "   "))
        line = f"  {tier:<32}{n:>6}   {marker} {TIER_DESCRIPTIONS.get(tier, '').splitlines()[0]}"
        print(painter.tier(tier, line))
    if counts["noise_floor"]:
        line = f"  {'noise_floor (silent)':<32}{counts['noise_floor']:>6}"
        print(painter.dim(line))
    if counts["skipped"]:
        line = f"  {'skipped':<32}{counts['skipped']:>6}"
        print(painter.dim(line))

    print()
    flagged_line = (
        f"  Pairs flagged at score >= {threshold:.2f}: "
        f"{n_flagged} of {n_evaluated}"
    )
    if n_flagged:
        print(painter.bold(flagged_line))
    else:
        print(flagged_line)


def _render_profile_catalog(painter: _Painter) -> None:
    """Print a table of available task profiles for ``--list-profiles``."""
    print(painter.bold("Available task profiles"))
    print("-----------------------")
    name_w = max(len(n) for n in TASK_PROFILES) + 2
    for name, profile in TASK_PROFILES.items():
        print(f"  {painter.bold(name.ljust(name_w))}  {profile.description}")


# ---------------------------------------------------------------------------
# Scanning driver
# ---------------------------------------------------------------------------


def _resolve_pair_ids(
    *, pairs: list[PairSpec], id_kind: str, config: Config,
) -> list[_ScannedPair]:
    """Resolve PDB IDs to UniProt accessions if needed.

    Returns one :class:`_ScannedPair` per input pair. PDBs that don't
    resolve to any accessions are marked ``skip_reason`` and the report
    is left None.
    """
    if id_kind == "accession":
        return [
            _ScannedPair(
                spec=p,
                test_resolved=p.test_ids,
                comparison_resolved=p.comparison_ids,
                report=None,
            )
            for p in pairs
        ]
    if id_kind != "pdb":
        raise SystemExit(f"Unknown --id-kind: {id_kind!r}")
    all_pdbs: set[str] = set()
    for p in pairs:
        all_pdbs.update(p.test_ids)
        all_pdbs.update(p.comparison_ids)
    resolution = _resolve_pdb_set(config, all_pdbs)
    out: list[_ScannedPair] = []
    for p in pairs:
        test_resolved: list[str] = []
        for pdb in p.test_ids:
            test_resolved.extend(resolution.get(pdb.upper(), []))
        comp_resolved: list[str] = []
        for pdb in p.comparison_ids:
            comp_resolved.extend(resolution.get(pdb.upper(), []))
        skip_reason: str | None = None
        if not test_resolved or not comp_resolved:
            missing_test = [p_ for p_ in p.test_ids if not resolution.get(p_.upper())]
            missing_comp = [p_ for p_ in p.comparison_ids if not resolution.get(p_.upper())]
            if missing_test and missing_comp:
                skip_reason = f"PDB unresolved: {missing_test[0]}, {missing_comp[0]}"
            elif missing_test:
                skip_reason = f"PDB unresolved: {missing_test[0]}"
            elif missing_comp:
                skip_reason = f"PDB unresolved: {missing_comp[0]}"
            else:
                skip_reason = "no UniProt accessions for this PDB"
        out.append(_ScannedPair(
            spec=p,
            test_resolved=sorted(set(test_resolved)),
            comparison_resolved=sorted(set(comp_resolved)),
            report=None,
            skip_reason=skip_reason,
        ))
    return out


# ---------------------------------------------------------------------------
# Live-progress tracker for the scan loop
# ---------------------------------------------------------------------------


# Short tokens per axis-discovery function. Order matches the runner's
# pipeline so the on-screen progression makes intuitive sense.
_AXIS_TOKENS: dict[str, str] = {
    "shared_ortholog_membership":     "orth",
    "shared_pathway_membership":      "path",
    "shared_interaction_network":     "intx",
    "shared_motif_domain_family":     "fam ",
    "shared_structural_classification": "stru",
    "shared_function_class":          "func",
    "shared_domain_architecture":     "arch",
}
_AXIS_COUNT = len(_AXIS_TOKENS)

# Number of recent pairs to average for the rolling ETA. Small enough that
# the estimate converges quickly to the warm-cache rate after the cold
# first pair.
_ETA_WINDOW = 3

# Spinner characters cycled while a pair is in flight. The simple
# four-frame `|/-\` works in every Windows console code page.
_SPINNER_FRAMES = "|/-\\"

# Heartbeat tick rate (seconds). Fast enough that the spinner looks alive,
# slow enough to keep refresh cost negligible.
_HEARTBEAT_INTERVAL = 0.25

# Maximum width we render. Anything beyond this gets truncated so we don't
# wrap on narrow terminals.
_PROGRESS_MAX_WIDTH = 110


def _fmt_duration(seconds: float) -> str:
    """Format a duration as ``Xs`` / ``XmYs`` / ``XhYm`` for display."""
    if seconds < 0:
        seconds = 0.0
    if seconds < 60:
        return f"{seconds:.0f}s"
    if seconds < 3600:
        m, s = divmod(int(round(seconds)), 60)
        return f"{m}m{s:02d}s"
    h, rem = divmod(int(round(seconds)), 3600)
    m, _ = divmod(rem, 60)
    return f"{h}h{m:02d}m"


@dataclass
class _ScanProgress:
    """Heartbeat + per-axis live progress for the scan loop.

    On a TTY the current pair is rendered with ``\\r`` so the same line
    updates in place as each axis finishes; a background daemon thread
    refreshes the spinner so even a 60-second cold-cache axis still shows
    activity. When stdout isn't a TTY (redirected to file, piped) we
    print one final line per pair after it completes — no in-place
    updates that would litter a log.
    """
    painter: _Painter
    n_pairs: int
    tty: bool
    width: int = _PROGRESS_MAX_WIDTH

    # ---- internal state, guarded by ``_lock`` -----------------------
    _lock: threading.Lock = field(default_factory=threading.Lock)
    _stop: threading.Event = field(default_factory=threading.Event)
    _thread: threading.Thread | None = None
    _tick: int = 0
    # Per-pair state
    _idx: int = 0
    _label: str = ""
    _t_pair_start: float = 0.0
    _tokens_done: list[str] = field(default_factory=list)
    _tokens_fired: list[bool] = field(default_factory=list)
    _current_axis_token: str = ""
    _line_painted: bool = False
    # Cumulative state
    _completed_times: list[float] = field(default_factory=list)

    # ----- thread lifecycle ------------------------------------------

    def start(self) -> None:
        """Start the heartbeat thread (TTY only)."""
        if not self.tty:
            return
        self._thread = threading.Thread(target=self._heartbeat_loop, daemon=True)
        self._thread.start()

    def stop(self) -> None:
        """Signal the heartbeat thread to exit and wait briefly."""
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=1.0)
            self._thread = None

    def _heartbeat_loop(self) -> None:
        while not self._stop.wait(_HEARTBEAT_INTERVAL):
            with self._lock:
                self._tick += 1
                if self._label and self.tty:
                    self._render_in_progress_locked()

    # ----- callbacks fired from the scan thread ----------------------

    def start_pair(self, idx: int, label: str) -> None:
        with self._lock:
            self._idx = idx
            self._label = label
            self._tokens_done = []
            self._tokens_fired = []
            self._current_axis_token = ""
            self._t_pair_start = time.monotonic()
            self._line_painted = False
            if self.tty:
                self._render_in_progress_locked()

    def on_axis_start(self, fn_name: str) -> None:
        token = _AXIS_TOKENS.get(fn_name, fn_name[:4])
        with self._lock:
            self._current_axis_token = token
            if self.tty:
                self._render_in_progress_locked()

    def on_axis_done(self, fn_name: str, result: dict) -> None:
        token = _AXIS_TOKENS.get(fn_name, fn_name[:4])
        with self._lock:
            self._tokens_done.append(token)
            self._tokens_fired.append(bool(result))
            self._current_axis_token = ""
            if self.tty:
                self._render_in_progress_locked()

    def end_pair(self, *, score: float | None, tier: str | None) -> None:
        with self._lock:
            elapsed = time.monotonic() - self._t_pair_start
            self._completed_times.append(elapsed)
            if self.tty:
                self._render_final_locked(elapsed=elapsed, score=score, tier=tier)
                sys.stdout.write("\n")
                sys.stdout.flush()
            else:
                # Non-TTY: emit a single clean line per pair.
                sys.stdout.write(
                    self._compose_line(
                        in_progress=False, elapsed=elapsed,
                        score=score, tier=tier,
                    ) + "\n"
                )
                sys.stdout.flush()
            # Reset per-pair state so the heartbeat thread stops painting
            # this pair after it's done.
            self._label = ""
            self._line_painted = False

    def skipped(self, idx: int, label: str, reason: str) -> None:
        """Render a single line for a skipped pair."""
        with self._lock:
            line = (
                f"  [{idx}/{self.n_pairs}]  {_truncate(label, 32):<32}  "
                f"{self.painter.dim('skipped: ' + reason)}"
            )
            sys.stdout.write(line + "\n")
            sys.stdout.flush()

    def announce_header(self, profile_name: str) -> None:
        title = f"Scanning {self.n_pairs} pair{'s' if self.n_pairs != 1 else ''} ('{profile_name}' profile)..."
        print(self.painter.bold(title))
        if self.tty and self.n_pairs > 1:
            print(self.painter.dim(
                "  (first pair pays a cold-cache cost; ETA recalibrates after each pair)"
            ))
        print()

    def announce_footer(self) -> None:
        total = sum(self._completed_times)
        if total <= 0:
            return
        n_done = len(self._completed_times)
        line = f"  Total scan time: {_fmt_duration(total)}"
        if n_done > 1:
            avg = total / n_done
            line += f"   (avg {_fmt_duration(avg)}/pair)"
        print()
        print(self.painter.dim(line))
        print()

    # ----- rendering -------------------------------------------------

    def _render_in_progress_locked(self) -> None:
        if not self.tty:
            return
        line = self._compose_line(in_progress=True, elapsed=None, score=None, tier=None)
        sys.stdout.write("\r" + line)
        sys.stdout.flush()
        self._line_painted = True

    def _render_final_locked(
        self, *, elapsed: float, score: float | None, tier: str | None,
    ) -> None:
        if not self.tty:
            return
        line = self._compose_line(
            in_progress=False, elapsed=elapsed, score=score, tier=tier,
        )
        sys.stdout.write("\r" + line)
        sys.stdout.flush()

    def _compose_line(
        self, *, in_progress: bool, elapsed: float | None,
        score: float | None, tier: str | None,
    ) -> str:
        # Prefix: [n/N] label                          (label left-padded)
        label_width = 32
        prefix = (
            f"  [{self._idx}/{self.n_pairs}]  "
            f"{_truncate(self._label, label_width):<{label_width}}  "
        )
        # Axis-token strip: 7 tokens, each 4 chars + separator. Total width
        # is fixed so subsequent pairs line up. Done axes are bold or dim
        # depending on whether they fired; the active axis (if any) gets a
        # spinner; pending axes show as `....`.
        token_cells: list[str] = []
        for i, token in enumerate(self._tokens_done):
            fired = self._tokens_fired[i] if i < len(self._tokens_fired) else False
            token_cells.append(self.painter.bold(token) if fired else self.painter.dim(token))
        if in_progress and self._current_axis_token:
            spinner = _SPINNER_FRAMES[self._tick % len(_SPINNER_FRAMES)]
            token_cells.append(self.painter.bold(self._current_axis_token[:3] + spinner))
        for _ in range(_AXIS_COUNT - len(token_cells)):
            token_cells.append(self.painter.dim("...."))
        tokens_str = " ".join(token_cells)
        # Time + status
        if in_progress:
            t_now = time.monotonic() - self._t_pair_start
            tail = f" T+{t_now:>5.1f}s"
            # ETA from completed pairs.
            tail += self._eta_tail(after_this=False)
        else:
            tail = f" {elapsed:>5.1f}s"
            if score is not None and tier is not None and self._idx < self.n_pairs:
                # Inline severity marker so users see the result at a glance.
                marker = _TIER_STYLE.get(tier, (_Ansi.DIM, "   "))[1]
                tail += f"  {self.painter.tier(tier, marker)}"
            tail += self._eta_tail(after_this=True)
        line = prefix + tokens_str + tail
        # Pad to terminal width so leftover characters from a previous
        # in-place update get blanked out on the new render.
        try:
            cols = shutil.get_terminal_size((self.width, 24)).columns
        except Exception:
            cols = self.width
        cols = max(MIN_TERM_WIDTH, min(cols, _PROGRESS_MAX_WIDTH))
        visible_len = _ansi_visible_length(line)
        if visible_len < cols:
            line = line + " " * (cols - visible_len)
        return line

    def _eta_tail(self, *, after_this: bool) -> str:
        # Use the most recent _ETA_WINDOW completed pairs so the cold-
        # cache first pair doesn't poison the estimate forever.
        if not self._completed_times:
            return ""
        idx = self._idx
        remaining = self.n_pairs - (idx if after_this else (idx - 1))
        if remaining <= 0:
            return ""
        window = self._completed_times[-_ETA_WINDOW:]
        avg = sum(window) / len(window)
        eta = avg * remaining
        return f"   eta {_fmt_duration(eta)}"


def _ansi_visible_length(text: str) -> int:
    """Length of ``text`` ignoring ANSI escape sequences."""
    out = 0
    i = 0
    n = len(text)
    while i < n:
        c = text[i]
        if c == "\x1b" and i + 1 < n and text[i + 1] == "[":
            # Skip CSI sequence: ESC [ ... letter
            j = i + 2
            while j < n and not ("@" <= text[j] <= "~"):
                j += 1
            i = j + 1
            continue
        out += 1
        i += 1
    return out


def _scan(
    *, pairs: list[_ScannedPair], config: Config, profile_name: str,
    progress: _ScanProgress | None = None,
) -> None:
    """Run the overlap engine on each pair, attaching reports in place.

    If ``progress`` is provided, it's notified at every pair boundary and
    every axis boundary so the CLI can show a heartbeat / ETA in the
    terminal while the scan runs.
    """
    for idx, p in enumerate(pairs, start=1):
        label = p.spec.label
        if p.skip_reason is not None:
            if progress is not None:
                progress.skipped(idx, label, p.skip_reason)
            continue
        if not p.test_resolved or not p.comparison_resolved:
            p.skip_reason = "empty after normalization"
            if progress is not None:
                progress.skipped(idx, label, p.skip_reason)
            continue
        if progress is not None:
            progress.start_pair(idx, label)
        p.report = discover_overlap(
            config,
            p.test_resolved,
            p.comparison_resolved,
            task_profile=profile_name,
            on_axis_start=(progress.on_axis_start if progress else None),
            on_axis_done=(progress.on_axis_done if progress else None),
        )
        if progress is not None:
            score = p.report.composite_score()
            tier = p.report.severity_tier()
            progress.end_pair(score=score, tier=tier)


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------


def _write_json_report(
    *, path: Path, pairs: list[_ScannedPair], profile_name: str, threshold: float,
) -> None:
    payload = {
        "profile": profile_name,
        "threshold": threshold,
        "pair_count": len(pairs),
        "pairs": [
            {
                "label": p.spec.label,
                "row_number": p.spec.row_number,
                "test_input": p.spec.test_ids,
                "comparison_input": p.spec.comparison_ids,
                "test_resolved": p.test_resolved,
                "comparison_resolved": p.comparison_resolved,
                "skip_reason": p.skip_reason,
                "report": p.report.to_dict() if p.report is not None else None,
            }
            for p in pairs
        ],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _write_csv_report(
    *, path: Path, pairs: list[_ScannedPair], threshold: float,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "label", "score", "severity_tier", "flagged",
            "test_resolved", "comparison_resolved",
            "skip_reason",
        ])
        for p in pairs:
            if p.report is None:
                writer.writerow([
                    p.spec.label, "", "", "",
                    ";".join(p.test_resolved or p.spec.test_ids),
                    ";".join(p.comparison_resolved or p.spec.comparison_ids),
                    p.skip_reason or "",
                ])
                continue
            score = p.report.composite_score()
            tier = p.report.severity_tier()
            flagged = score >= threshold or tier == "identity"
            writer.writerow([
                p.spec.label,
                f"{score:.4f}",
                tier,
                "yes" if flagged else "no",
                ";".join(p.test_resolved),
                ";".join(p.comparison_resolved),
                "",
            ])


# ---------------------------------------------------------------------------
# Argparse + main
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="proteosphere pairwise-checker",
        description=(
            "Score per-pair biological-similarity overlap between protein "
            "accession sets. Each row of the input is one comparison; the "
            "output is a sorted, color-coded summary and an optional JSON / "
            "CSV report. (Also reachable under the legacy name `overlap-scan`.)"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  proteosphere pairwise-checker --input pairs.csv\n"
            "  proteosphere pairwise-checker --input pairs.csv --profile structure_prediction --out report.json\n"
            "  proteosphere pairwise-checker --input pdbs.csv --id-kind pdb --threshold 0.20\n"
            "  proteosphere pairwise-checker --list-profiles\n"
        ),
    )
    parser.add_argument(
        "--input", "-i", type=Path,
        help="Path to the input file (.csv .tsv .json .xlsx).",
    )
    parser.add_argument(
        "--profile", "-p", default=DEFAULT_PROFILE,
        choices=sorted(TASK_PROFILES),
        help=f"Task profile to apply (default: {DEFAULT_PROFILE}).",
    )
    parser.add_argument(
        "--out", "-o", type=Path, default=None,
        help="Write a full JSON report (one entry per pair) to this path.",
    )
    parser.add_argument(
        "--csv-out", type=Path, default=None,
        help="Write a per-pair CSV summary to this path.",
    )
    parser.add_argument(
        "--id-kind", default="accession",
        choices=("accession", "pdb"),
        help=(
            "How to interpret the ID columns. 'accession' (default) treats "
            "values as UniProt accessions; 'pdb' resolves PDB IDs to UniProt "
            "via the structure_units table."
        ),
    )
    parser.add_argument(
        "--threshold", "-t", type=float, default=DEFAULT_THRESHOLD,
        help=(
            f"Composite-score threshold for the 'flagged' marker "
            f"(default: {DEFAULT_THRESHOLD:.2f})."
        ),
    )
    parser.add_argument(
        "--quiet", "-q", action="store_true",
        help="Skip the per-pair table; only print the summary.",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Print every fired axis and weight under each pair row.",
    )
    parser.add_argument(
        "--no-color", action="store_true",
        help="Disable ANSI color output (useful for redirecting to files).",
    )
    parser.add_argument(
        "--list-profiles", action="store_true",
        help="Print the available task profiles and exit.",
    )
    return parser


def main(argv: list[str] | None = None, *, config: Config | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    use_color = not args.no_color and _supports_color()
    if use_color:
        _enable_ansi_on_windows()
    painter = _Painter(enabled=use_color)

    if args.list_profiles:
        _render_profile_catalog(painter)
        return 0

    if args.input is None:
        parser.error("--input is required (or pass --list-profiles).")

    input_path: Path = args.input.expanduser().resolve()
    if not input_path.is_file():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        return 2

    try:
        pairs = load_pairs(input_path)
    except (ValueError, json.JSONDecodeError) as e:
        print(f"ERROR: failed to read {input_path}:\n  {e}", file=sys.stderr)
        return 2
    if not pairs:
        print(f"WARNING: no pairs found in {input_path}.", file=sys.stderr)
        return 0

    config = resolve_config_or_friendly_exit(config, painter=painter)

    width = _terminal_width()
    _render_header(
        painter=painter,
        input_path=input_path,
        n_pairs=len(pairs),
        profile=args.profile,
        threshold=args.threshold,
        warehouse_root=config.warehouse_root,
        id_kind=args.id_kind,
        width=width,
    )

    scanned = _resolve_pair_ids(pairs=pairs, id_kind=args.id_kind, config=config)

    # Live-progress tracker: heartbeat in the terminal, one final line per
    # pair when redirected to a file. Skipped with --quiet so log-style
    # invocations stay compact.
    if args.quiet:
        progress = None
    else:
        # Whether to do in-place \r updates is independent of color: a
        # user piping to `less -R` may pass --no-color but still want the
        # spinner. Repainting needs an actual interactive terminal.
        is_tty = sys.stdout.isatty()
        progress = _ScanProgress(
            painter=painter,
            n_pairs=len(scanned),
            tty=is_tty,
        )
        progress.announce_header(args.profile)
        progress.start()
    try:
        _scan(
            pairs=scanned, config=config,
            profile_name=args.profile, progress=progress,
        )
    finally:
        if progress is not None:
            progress.stop()
            progress.announce_footer()

    if not args.quiet:
        _render_pair_table(
            painter=painter,
            pairs=scanned,
            threshold=args.threshold,
            width=width,
            verbose=args.verbose,
        )
    _render_summary(painter=painter, pairs=scanned, threshold=args.threshold)

    n_flagged = sum(
        1 for p in scanned
        if p.report is not None and (
            p.report.composite_score() >= args.threshold
            or p.report.severity_tier() == "identity"
        )
    )

    if args.out is not None:
        out_path: Path = args.out.expanduser().resolve()
        _write_json_report(
            path=out_path, pairs=scanned,
            profile_name=args.profile, threshold=args.threshold,
        )
        print()
        print(f"  Wrote JSON report -> {out_path}")
    if args.csv_out is not None:
        csv_path: Path = args.csv_out.expanduser().resolve()
        _write_csv_report(
            path=csv_path, pairs=scanned, threshold=args.threshold,
        )
        print(f"  Wrote CSV summary -> {csv_path}")
    sys.stdout.flush()
    return 1 if n_flagged else 0


if __name__ == "__main__":
    raise SystemExit(main())
